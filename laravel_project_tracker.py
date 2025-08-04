# laravel_project_tracker.py

"""
Laravel Project Tracker
Scans Laravel project folders and generates a detailed Excel file with cross-linked routes, controllers, models, views, JS, configs, migrations, tests, and more.
"""

# ========== 🧩 IMPORTS ==========
import os
import re
import sys
import uuid
import json
import logging
import datetime
import platform
import io
import hashlib
import multiprocessing
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

from logging.handlers import RotatingFileHandler

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.comments import Comment

from openpyxl.utils import quote_sheetname

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    os.environ["PYTHONIOENCODING"] = "utf-8"

# ========== 🏁 GLOBAL FLAGS ==========
LOG_TO_CONSOLE = True
SKIP_DUPLICATE_COMPONENT_ID = True
OUTPUT_FILE_NAME = "laravel_project_tracker.xlsx"
MAX_WORKERS = min(16, multiprocessing.cpu_count() * 2)
HIGHLIGHT_EMPTY_CELLS = True
DEBUG_PRINT_MINIMAL = False

# ========== 🗂️ PATH SETUP ==========
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_PROJECT_ROOT = BASE_DIR.parent
CONFIG_PATH = BASE_DIR / "config.json"
LOG_FILE = BASE_DIR / "tracker_errors.log"
EXCEL_OUTPUT_PATH = BASE_DIR / OUTPUT_FILE_NAME

# Will be overridden by config.json if exists
PROJECT_ROOT = DEFAULT_PROJECT_ROOT

# ========== 🎨 COLOR SCHEME ==========
COLOR_SCHEME = {
    "Routes": "FFF2CC",       # Light Orange
    "Controllers": "E2EFDA",  # Light Green
    "Models": "DEEBF7",       # Light Blue
    "Views": "FCE4D6",        # Light Red
    "JavaScript": "EDEDED",   # Gray
    "Database": "D9D9D9"      # Dark Gray
}

# ========== 📋 SHEET ORDER ==========
SHEET_ORDER = [
    # 🔖 Overview
    "TOC",
    "Dashboard",

    # 🌐 Routing Layer
    "Web Routes",
    "API Routes",
    "Route References",

    # 🧠 Controllers & Views
    "Controllers",
    "Views",
    "Blade Hierarchy",
    "Livewire Components",

    # 📦 Models & DB
    "Models",
    "Model Relationships",
    "Migrations",
    "Database Schema",
    "Seeders & Factories",

    # 🧩 Frontend/Assets
    "JavaScript",

    # ⚙️ App Structure & Config
    "Middleware",
    "Service Providers",
    "Config Files",
    "Config Usage",
    "Environment Usage",

    # 🔐 Security & Validation
    "Validation Rules",
    "Authorization Map",
    "Security Audit",

    # 🧪 Testing & Exports
    "Tests",
    "Data Exporters",
    "Manual QA",

    # 📚 Dependencies
    "Dependencies",
    "Service Dependencies",

    # 📁 Filesystem Structure
    "Folder Structure",

    # 🔗 Cross-Reference
    "Events & Listeners",
    "Component Relationships",
    "Master Reference"
]

# ========== 🧠 GLOBAL STRUCTURES ==========
sheet_data = {}               # {sheet_name: [rows]}
component_index = {}          # {component_id: {sheet, row, name, file}}
sheet_meta = {}               # {sheet_name: {entries, last_modified}}

# Mappings for relationships
route_controller_map = defaultdict(list)
controller_model_map = defaultdict(list)
view_controller_map = defaultdict(set)
js_route_map = defaultdict(list)

# ========== ⚙️ CONFIGURATION LOADER ==========
def load_config():
    global PROJECT_ROOT, EXCEL_OUTPUT_PATH, MAX_WORKERS

    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                config = json.load(f)
                logging.info("[✓] Loaded config.json")

                if "project_root" in config:
                    PROJECT_ROOT = Path(config["project_root"]).resolve()
                if "output_file" in config:
                    EXCEL_OUTPUT_PATH = Path(config["output_file"]).resolve()
                if "max_workers" in config:
                    MAX_WORKERS = int(config["max_workers"])

                return config
    except Exception as e:
        print(f"[✗] Failed to load config.json: {e}")
        logging.warning(f"Failed to load config.json: {e}")

    return {}

# ========== 🧾 LOGGING SETUP ==========
def setup_logging():
    log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler = RotatingFileHandler(LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=2)
    file_handler.setFormatter(log_formatter)

    handlers = [file_handler]

    if LOG_TO_CONSOLE:
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setFormatter(log_formatter)
        handlers.append(stream_handler)

    logging.basicConfig(level=logging.INFO, handlers=handlers)
    logging.info(" Logging initialized")

# Call immediately for early errors
setup_logging()

# ========== 🔧 UTILITY FUNCTIONS ==========

def generate_component_id(prefix: str):
    """Generate a unique component ID with validation"""
    if not prefix or not isinstance(prefix, str):
        prefix = "GEN"
    prefix = prefix.upper()[:3]
    return f"{prefix}-{str(uuid.uuid4())[:8]}"

def slugify(name: str):
    return re.sub(r"[^\w\-]", "_", name.strip().lower())

def get_current_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def clean_code_snippet(snippet: str, max_len=100):
    return snippet.strip().replace('\n', ' ')[:max_len]

def readable_file_size(bytes_value):
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_value < 1024.0:
            return f"{bytes_value:.1f} {unit}"
        bytes_value /= 1024.0
    return f"{bytes_value:.1f} TB"

def is_laravel_route_line(line):
    return any(kw in line for kw in ['Route::', '->group', '->middleware', '->name'])

def find_php_class_name(file_content):
    match = re.search(r'class\s+([A-Za-z0-9_]+)', file_content)
    return match.group(1) if match else ""

def extract_php_namespace(file_content):
    match = re.search(r'namespace\s+(.+?);', file_content)
    return match.group(1).strip() if match else ""

# ========== 🎨 EXCEL FORMAT HELPERS ==========

def get_header_fill():
    return PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

def get_bold_font():
    return Font(bold=True)

def get_center_alignment():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def get_border():
    return Border(
        left=Side(style='thin', color='999999'),
        right=Side(style='thin', color='999999'),
        top=Side(style='thin', color='999999'),
        bottom=Side(style='thin', color='999999')
    )

def write_cell(ws, row_idx, col_idx, value):
    cell = ws.cell(row=row_idx, column=col_idx, value=value)

    if HIGHLIGHT_EMPTY_CELLS and (value is None or value == ""):
        cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")

    return cell

def apply_auto_column_width(ws, headers, max_width=60):
    for idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(idx)
        max_len = len(str(header))

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass

        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)

def freeze_header(ws):
    ws.freeze_panes = ws['A2']

# ========== 🪄 BASE SHEET FORMATTER ==========

def apply_base_formatting(ws, headers, data_rows):
    header_fill = get_header_fill()
    bold_font = get_bold_font()
    align = get_center_alignment()
    border = get_border()

    # Apply headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = align
        cell.border = border

    # Write data
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, key in enumerate(headers, 1):
            val = row.get(key, "")
            cell = write_cell(ws, row_idx, col_idx, val)
            cell.border = border

    apply_auto_column_width(ws, headers)
    freeze_header(ws)

def clean_blade_placeholders(s):
    if not s:
        return ""
    return re.sub(r'\{\{.*?\}\}', '', s).strip().strip('/')

# ========== 🌐 ROUTE SCANNERS ==========

def scan_web_routes():
    logging.info(" Scanning routes/web.php ...")
    route_file = PROJECT_ROOT / "routes" / "web.php"
    results = []

    if not route_file.exists():
        logging.warning(" web.php not found.")
        return "Web Routes", []

    try:
        with open(route_file, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for idx, line in enumerate(lines):
            line = line.strip()
            if not is_laravel_route_line(line):
                continue

            entry = {
                "Component ID": generate_component_id("WR"),
                "Route Method": "",
                "URI": "",
                "Controller": "",
                "Route Name": "",
                "Middleware": "",
                "Defined At": f"{route_file.name}:{idx+1}",
                "Controller ID": ""
            }

            try:
                # Route method
                method_match = re.search(r'Route::(get|post|put|patch|delete|any)', line)
                if method_match:
                    entry["Route Method"] = method_match.group(1).upper()

                # URI
                uri_match = re.search(r'Route::[a-z]+\([\'"](.+?)[\'"]', line)
                if uri_match:
                    entry["URI"] = uri_match.group(1)

                # Controller
                ctrl_match = re.search(r'->(controller|uses|action)\([\'"](.+?)@(.+?)[\'"]\)', line)
                if ctrl_match:
                    entry["Controller"] = ctrl_match.group(2)
                else:
                    ctrl_inline = re.search(r'\[?([A-Za-z0-9_\\]+)@([a-zA-Z0-9_]+)\]?', line)
                    if ctrl_inline:
                        entry["Controller"] = ctrl_inline.group(1)

                # Route name
                name_match = re.search(r'->name\([\'"](.+?)[\'"]\)', line)
                if name_match:
                    entry["Route Name"] = name_match.group(1)

                # Middleware
                middleware_match = re.search(r'->middleware\((.+?)\)', line)
                if middleware_match:
                    entry["Middleware"] = middleware_match.group(1).replace("'", "").replace('"', '')

            except Exception as e:
                entry["Error"] = f"Parse error: {str(e)}"
                logging.warning(f" Route parse error @ web.php:{idx+1}: {e}")

            

            route_controller_map[entry["Controller"]].append(entry["Component ID"])
            results.append(entry)

        if DEBUG_PRINT_MINIMAL:
            print(f"✓ Scanned Web Routes: {len(results)} entries")

    except Exception as e:
        logging.error(f" Failed to scan web.php: {e}")

    return "Web Routes", results


def scan_api_routes():
    logging.info(" Scanning routes/api.php ...")
    route_file = PROJECT_ROOT / "routes" / "api.php"
    results = []

    if not route_file.exists():
        logging.warning(" api.php not found.")
        return "API Routes", []

    try:
        with open(route_file, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for idx, line in enumerate(lines):
            line = line.strip()
            if not is_laravel_route_line(line):
                continue

            entry = {
                "Component ID": generate_component_id("AR"),
                "Route Method": "",
                "URI": "",
                "Controller": "",
                "Route Name": "",
                "Middleware": "",
                "Defined At": f"{route_file.name}:{idx+1}",
                "Controller ID": ""
            }

            try:
                method_match = re.search(r'Route::(get|post|put|patch|delete|any)', line)
                if method_match:
                    entry["Route Method"] = method_match.group(1).upper()

                uri_match = re.search(r'Route::[a-z]+\([\'"](.+?)[\'"]', line)
                if uri_match:
                    entry["URI"] = uri_match.group(1)

                ctrl_match = re.search(r'->(controller|uses|action)\([\'"](.+?)@(.+?)[\'"]\)', line)
                if ctrl_match:
                    entry["Controller"] = ctrl_match.group(2)
                else:
                    ctrl_inline = re.search(r'\[?([A-Za-z0-9_\\]+)@([a-zA-Z0-9_]+)\]?', line)
                    if ctrl_inline:
                        entry["Controller"] = ctrl_inline.group(1)

                name_match = re.search(r'->name\([\'"](.+?)[\'"]\)', line)
                if name_match:
                    entry["Route Name"] = name_match.group(1)

                middleware_match = re.search(r'->middleware\((.+?)\)', line)
                if middleware_match:
                    entry["Middleware"] = middleware_match.group(1).replace("'", "").replace('"', '')

            except Exception as e:
                entry["Error"] = f"Parse error: {str(e)}"
                logging.warning(f" Route parse error @ api.php:{idx+1}: {e}")

            

            route_controller_map[entry["Controller"]].append(entry["Component ID"])
            results.append(entry)

        if DEBUG_PRINT_MINIMAL:
            print(f"✓ Scanned API Routes: {len(results)} entries")

    except Exception as e:
        logging.error(f" Failed to scan api.php: {e}")

    return "API Routes", results

# ========== 🧭 CONTROLLER SCANNER ==========

def scan_controllers():
    logging.info(" Scanning Controllers ...")
    controller_dir = PROJECT_ROOT / "app" / "Http" / "Controllers"
    results = []

    if not controller_dir.exists():
        logging.warning(" No Controllers found.")
        return "Controllers", []

    # First build a list of all model names for more accurate detection
    model_names = set()
    model_dir = PROJECT_ROOT / "app" / "Models"
    if model_dir.exists():
        for root, _, files in os.walk(model_dir):
            for file in files:
                if file.endswith(".php"):
                    model_names.add(file.replace(".php", ""))

    for root, _, files in os.walk(controller_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                namespace = extract_php_namespace(content)
                full_class = f"{namespace}\\{class_name}" if namespace else class_name

                component_id = generate_component_id("C")

                entry = {
                    "Component ID": component_id,
                    "Controller": class_name,
                    "Namespace": namespace,
                    "Full Class": full_class,
                    "Defined At": str(path.relative_to(PROJECT_ROOT)),
                    "Uses Models": "",
                    "Used By Routes": "",
                    "Uses Views": ""  # Added new field for view relationships
                }

                # Improved model detection with namespace awareness
                used_models = set()
                if model_names:  # Only if we found models
                    # Find model references with and without namespace
                    model_refs = re.findall(r'\b([A-Z][a-zA-Z0-9_]*)::', content)
                    model_refs += re.findall(r'\bnew\s+([A-Z][a-zA-Z0-9_]*)\(', content)
                    model_refs += re.findall(r'\b([A-Z][a-zA-Z0-9_]*)\s*\(', content)  # Method calls
                    
                    for ref in set(model_refs):
                        if ref in model_names:
                            used_models.add(ref)
                            if ref not in controller_model_map[class_name]:
                                controller_model_map[class_name].append(ref)

                if used_models:
                    entry["Uses Models"] = ", ".join(sorted(used_models))

                # Link route back to controller
                route_ids = route_controller_map.get(class_name, [])
                if route_ids:
                    entry["Used By Routes"] = ", ".join(route_ids)

                # Detect view references in controller methods
                view_refs = set()
                view_matches = re.findall(r'view\([\'"](.+?)[\'"]', content)
                view_matches += re.findall(r'View::make\([\'"](.+?)[\'"]', content)
                for view in view_matches:
                    view_refs.add(view.split('.')[0])  # Get base view name
                
                if view_refs:
                    entry["Uses Views"] = ", ".join(sorted(view_refs))
                    for view_name in view_refs:
                        if view_name not in view_controller_map:
                            view_controller_map[view_name] = set()
                        view_controller_map[view_name].add(class_name)

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Controller": file,
                    "Namespace": "",
                    "Full Class": "",
                    "Defined At": str(path),
                    "Uses Models": "",
                    "Used By Routes": "",
                    "Uses Views": "",
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning controller {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Controllers: {len(results)} entries")

    return "Controllers", results

# ========== 🧩 MODEL SCANNER ==========

def scan_models():
    logging.info(" Scanning Models ...")
    model_dir = PROJECT_ROOT / "app" / "Models"
    results = []

    if not model_dir.exists():
        logging.warning(" No Models found.")
        return "Models", []

    for root, _, files in os.walk(model_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                namespace = extract_php_namespace(content)
                full_class = f"{namespace}\\{class_name}" if namespace else class_name

                component_id = generate_component_id("M")

                entry = {
                    "Component ID": component_id,
                    "Model": class_name,
                    "Namespace": namespace,
                    "Full Class": full_class,
                    "Defined At": str(path.relative_to(PROJECT_ROOT)),
                    "Used In Controllers": "",
                }

                # Reverse-link model to controllers
                used_in = []
                for ctrl, models in controller_model_map.items():
                    if class_name in models:
                        used_in.append(ctrl)

                if used_in:
                    entry["Used In Controllers"] = ", ".join(sorted(set(used_in)))

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Model": file,
                    "Namespace": "",
                    "Full Class": "",
                    "Defined At": str(path),
                    "Used In Controllers": "",
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning model {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Models: {len(results)} entries")

    return "Models", results

# ========== 🖼️ VIEW SCANNER ==========

def scan_views():
    logging.info(" Scanning Views ...")
    views_dir = PROJECT_ROOT / "resources" / "views"
    results = []

    if not views_dir.exists():
        logging.warning(" Views folder missing.")
        return "Views", []

    # Enhanced patterns to detect various route references
    route_patterns = [
        r"route\(['\"](.+?)['\"]\)",       # route() helper
        r"@route\(['\"](.+?)['\"]\)",      # Blade directive
        r"url\(['\"](.+?)['\"]\)",         # url() helper
        r"action\(['\"](.+?)['\"]\)",      # Form actions
        r"redirect\(\)->route\(['\"](.+?)['\"]\)",  # Redirects
        r"['\"]/(.+?)['\"]",               # Raw URLs
        r"window\.route\(['\"](.+?)['\"]\)" # Ziggy.js
    ]

    for root, _, files in os.walk(views_dir):
        for file in files:
            if not file.endswith(".blade.php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")
            view_name = str(path.relative_to(views_dir)).replace(".blade.php", "").replace("\\", "/").replace("/", ".")

            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                component_id = generate_component_id("V")

                # Initialize entry with basic info
                entry = {
                    "Component ID": component_id,
                    "View": view_name,
                    "File": rel_path,
                    "Extends": "",
                    "Includes": "",
                    "Used By Controllers": "",
                    "Route References": "",
                    "Inline JavaScript": "No",
                    "Form Actions": ""
                }

                # 1. Detect view inheritance
                extends_matches = re.findall(r'@extends\([\'"](.+?)[\'"]\)', content)
                if extends_matches:
                    entry["Extends"] = ", ".join(sorted(set(extends_matches)))

                # 2. Detect included components
                includes_matches = re.findall(r'@include\([\'"](.+?)[\'"]\)', content)
                if includes_matches:
                    entry["Includes"] = ", ".join(sorted(set(includes_matches)))

                # 3. Detect all route references
                route_references = set()
                for pattern in route_patterns:
                    matches = re.findall(pattern, content)
                    for match in matches:
                        if isinstance(match, tuple):  # For patterns with groups
                            match = next(m for m in match if m)
                        route_references.add(match)
                
                if route_references:
                    entry["Route References"] = ", ".join(sorted(route_references))

                # 4. Detect inline JavaScript
                if '<script>' in content:
                    entry["Inline JavaScript"] = "Yes"

                # 5. Detect form actions
                form_actions = set()
                form_matches = re.findall(r'<form\s+[^>]*action=[\'"](.+?)[\'"]', content, re.IGNORECASE)
                form_actions.update(form_matches)
                if form_actions:
                    entry["Form Actions"] = ", ".join(sorted(form_actions))

                # Link back to controller if known
                if view_name not in view_controller_map:
                    view_controller_map[view_name] = set()

                controllers_using = view_controller_map[view_name]
                if controllers_using:
                    entry["Used By Controllers"] = ", ".join(sorted(controllers_using))

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "View": file,
                    "File": rel_path,
                    "Extends": "",
                    "Includes": "",
                    "Used By Controllers": "",
                    "Route References": "",
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning view {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Views: {len(results)} entries (with route references)")

    return "Views", results

def generate_route_reference_report():
    """Generates a report of all route references across the project"""
    logging.info(" Generating Route Reference Report ...")
    report = []

    # 1. Get all defined routes
    route_lookup = {}
    for route_type in ["Web Routes", "API Routes"]:
        for route in sheet_data.get(route_type, []):
            if route.get("Component ID") == "ERR":
                continue

            route_name = route.get("Route Name", "")
            uri = route.get("URI", "")
            component_id = route.get("Component ID", "")

            if route_name:
                route_lookup[route_name] = {
                    "Route Name": route_name,
                    "Route URI": uri,
                    "Route Type": route_type,
                    "Route ID": component_id,
                }

            if uri:
                route_lookup[uri] = {
                    "Route Name": route_name,
                    "Route URI": uri,
                    "Route Type": route_type,
                    "Route ID": component_id,
                }

    # 2. Define a helper to add entries
    def add_reference(ref_source, ref_value, file, component_label):
        ref_value = ref_value.strip().strip("/")
        matched = route_lookup.get(ref_value)

        # Try fuzzy URI fallback if not matched
        if not matched:
            for k, v in route_lookup.items():
                if ref_value == k or ref_value == k.strip("/"):
                    matched = v
                    break
                if ref_value in k or k in ref_value:
                    matched = v
                    break

        report.append({
            "Reference": ref_value,
            "Route Name": matched.get("Route Name", "") if matched else "",
            "Route URI": matched.get("Route URI", "") if matched else "",
            "Route Type": matched.get("Route Type", "") if matched else "",
            "Route ID": matched.get("Route ID", "") if matched else "",
            "Matches": "Yes" if matched else "No",
            "Referenced In": ref_source,
            "Component": component_label,
            "File": file
        })

    # 3. References from Views
    for view in sheet_data.get("Views", []):
        refs = view.get("Route References", "")
        if not refs:
            continue
        for ref in refs.split(", "):
            add_reference("View", ref.strip(), view["File"], view["View"])

    # 4. References from JavaScript
    for js in sheet_data.get("JavaScript", []):
        uris = js.get("Fetch/Route URIs", "")
        if not uris:
            continue
        for uri in uris.split(", "):
            add_reference("JavaScript", uri.strip(), js["File"], js["File"])

    # 5. References from Controllers
    for ctrl in sheet_data.get("Controllers", []):
        refs = ctrl.get("Used Routes", "")
        if not refs:
            continue
        for ref in refs.split(", "):
            add_reference("Controller", ref.strip(), ctrl["Defined At"], ctrl["Controller"])

    # 6. References from Tests
    for test in sheet_data.get("Tests", []):
        refs = test.get("Tested URIs", "")
        if not refs:
            continue
        for ref in refs.split(", "):
            add_reference("Test", ref.strip(), test["Defined At"], test["Test Class"])

    return "Route References", report

def generate_blade_hierarchy():
    logging.info(" Generating Blade Hierarchy Map ...")
    results = []

    for view in sheet_data.get("Views", []):
        view_name = view.get("View")
        file = view.get("File")

        # Extends
        extends = view.get("Extends", "")
        if extends:
            for ext in extends.split(","):
                ext = ext.strip()
                if ext:
                    results.append({
                        "View": view_name,
                        "File": file,
                        "Type": "Extends",
                        "Related View": ext
                    })

        # Includes
        includes = view.get("Includes", "")
        if includes:
            for inc in includes.split(","):
                inc = inc.strip()
                if inc:
                    results.append({
                        "View": view_name,
                        "File": file,
                        "Type": "Includes",
                        "Related View": inc
                    })

    return "Blade Hierarchy", results

# ========== ⚙️ JAVASCRIPT SCANNER ==========

def scan_javascript():
    logging.info(" Scanning JavaScript ...")
    js_dir = PROJECT_ROOT / "resources" / "js"
    results = []
    route_uri_map = {}

    for route_sheet in ["Web Routes", "API Routes"]:
        for route in sheet_data.get(route_sheet, []):
            uri = route.get("URI", "")
            route_uri_map[uri] = (route.get("Component ID"), route.get("Controller"))

    if not js_dir.exists():
        logging.warning(" JS folder missing.")
        return "JavaScript", []

    for root, _, files in os.walk(js_dir):
        for file in files:
            if not file.endswith(".js"):
                continue

            path = Path(root) / file
            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()

                component_id = generate_component_id("JS")
                rel_path = str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")

                entry = {
                    "Component ID": component_id,
                    "File": rel_path,
                    "Uses URI": "",
                    "Fetch/Route URIs": "",
                    "Related Controller": "",
                }

                uris_used = set(re.findall(r"['\"]\/([a-zA-Z0-9_\-/]+)['\"]", content))
                matched_ids = []
                matched_controllers = []

                for uri in uris_used:
                    for route_uri, (rid, ctrl) in route_uri_map.items():
                        if uri in route_uri:
                            matched_ids.append(rid)
                            matched_controllers.append(ctrl)

                if matched_ids:
                    entry["Uses URI"] = ", ".join(sorted(set(matched_ids)))
                if uris_used:
                    entry["Fetch/Route URIs"] = ", ".join(sorted(set(uris_used)))
                if matched_controllers:
                    entry["Related Controller"] = ", ".join(sorted(set(matched_controllers)))

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "File": str(path),
                    "Uses URI": "",
                    "Fetch/Route URIs": "",
                    "Related Controller": "",
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning JS file {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned JavaScript: {len(results)} entries")

    return "JavaScript", results


# ========== 🔐 MIDDLEWARE SCANNER ==========

def scan_middleware():
    logging.info(" Scanning Middleware ...")
    mw_dir = PROJECT_ROOT / "app" / "Http" / "Middleware"
    results = []

    if not mw_dir.exists():
        logging.warning(" Middleware folder missing.")
        return "Middleware", []

    for root, _, files in os.walk(mw_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                namespace = extract_php_namespace(content)
                full_class = f"{namespace}\\{class_name}" if namespace else class_name

                component_id = generate_component_id("MW")

                entry = {
                    "Component ID": component_id,
                    "Middleware": class_name,
                    "Namespace": namespace,
                    "Defined At": rel_path,
                }

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Middleware": file,
                    "Namespace": "",
                    "Defined At": rel_path,
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning middleware {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Middleware: {len(results)} entries")

    return "Middleware", results

# ========== 📦 MIGRATION SCANNER ==========

def scan_migrations():
    logging.info(" Scanning Migrations ...")
    migrations_dir = PROJECT_ROOT / "database" / "migrations"
    results = []

    if not migrations_dir.exists():
        logging.warning(" Migrations folder missing.")
        return "Migrations", []

    for root, _, files in os.walk(migrations_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                table_match = re.search(r'Schema::create\([\'"](.+?)[\'"]', content)
                fields = re.findall(r'\$(table)->(string|integer|bigInteger|text|boolean|date|timestamp)\([\'"](.+?)[\'"]', content)

                component_id = generate_component_id("MIG")

                entry = {
                    "Component ID": component_id,
                    "Migration File": file,
                    "Defines Table": table_match.group(1) if table_match else "",
                    "Fields": ", ".join(f[2] for f in fields),
                    "Defined At": rel_path,
                }

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Migration File": file,
                    "Defines Table": "",
                    "Fields": "",
                    "Defined At": rel_path,
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning migration {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Migrations: {len(results)} entries")

    return "Migrations", results

def initialize_lookups():
    """Create and return empty lookup dictionaries"""
    return {
        'routes_by_uri': {},
        'routes_by_name': {},
        'controllers_by_name': {},
        'models_by_name': {},
        'views_by_name': {},
        'views_by_path': {}
    }

# ========== 🧱 DATABASE SCHEMA FLATTENER ==========

def scan_schema():
    logging.info(" Generating Database Schema (flattened from migrations)...")
    results = []

    for entry in sheet_data.get("Migrations", []):
        if entry.get("Component ID") == "ERR":
            continue
        fields = entry.get("Fields", "").split(",")
        for field in fields:
            field = field.strip()
            if not field:
                continue
            schema_id = generate_component_id("DBF")
            schema_entry = {
                "Component ID": schema_id,
                "Table": entry.get("Defines Table", ""),
                "Field": field,
                "Defined In": entry.get("Migration File", ""),
            }
            results.append(schema_entry)

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Generated Schema Fields: {len(results)} entries")

    return "Database Schema", results


# ========== ⚙️ CONFIG SCANNER ==========

def scan_config():
    logging.info(" Scanning Config Files ...")
    config_dir = PROJECT_ROOT / "config"
    results = []

    if not config_dir.exists():
        logging.warning(" Config folder missing.")
        return "Config Files", []

    for root, _, files in os.walk(config_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    content = f.read()

                defines = re.findall(r"'(.+?)'\s*=>", content)

                component_id = generate_component_id("CFG")

                entry = {
                    "Component ID": component_id,
                    "Config File": file,
                    "Defined Keys": ", ".join(defines[:20]),
                    "Defined At": rel_path,
                }

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Config File": file,
                    "Defined Keys": "",
                    "Defined At": rel_path,
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning config file {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Config Files: {len(results)} entries")

    return "Config Files", results


# ========== 🧩 SERVICE PROVIDER SCANNER ==========

def scan_service_providers():
    logging.info(" Scanning Service Providers ...")
    providers_dir = PROJECT_ROOT / "app" / "Providers"
    results = []

    if not providers_dir.exists():
        logging.warning(" Service Providers folder missing.")
        return "Service Providers", []

    for root, _, files in os.walk(providers_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                namespace = extract_php_namespace(content)
                full_class = f"{namespace}\\{class_name}" if namespace else class_name

                component_id = generate_component_id("PRV")

                entry = {
                    "Component ID": component_id,
                    "Provider Class": class_name,
                    "Namespace": namespace,
                    "Full Class": full_class,
                    "Defined At": rel_path,
                }

                

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Provider Class": file,
                    "Namespace": "",
                    "Full Class": "",
                    "Defined At": rel_path,
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning provider {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Service Providers: {len(results)} entries")

    return "Service Providers", results

# ========== 🧪 TEST SCANNER ==========

def scan_tests():
    logging.info(" Scanning Tests ...")
    test_dir = PROJECT_ROOT / "tests"
    results = []

    # Build route lookup (flattened)
    route_lookup = {}
    for route_sheet in ["Web Routes", "API Routes"]:
        for route in sheet_data.get(route_sheet, []):
            uri = route.get("URI", "").strip().strip("/")
            if uri:
                variations = {
                    uri,
                    "/" + uri,
                    uri + "/",
                    "/" + uri + "/"
                }
                for v in variations:
                    route_lookup[v] = {
                        "Component ID": route.get("Component ID"),
                        "Route Name": route.get("Route Name"),
                        "Controller": route.get("Controller"),
                        "Sheet": route_sheet
                    }

    if not test_dir.exists():
        logging.warning(" Tests folder missing.")
        return "Tests", []

    for root, _, files in os.walk(test_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                namespace = extract_php_namespace(content)
                test_methods = re.findall(r'public function (test[A-Za-z0-9_]+)\(', content)

                # 🔍 Extract URI usages in $this->get('/...'), $this->post('...'), etc.
                uris_called = re.findall(r'\$this->(get|post|put|delete|patch)\([\'"](.+?)[\'"]', content)
                tested_uris = set()
                tested_routes = set()

                for _, uri in uris_called:
                    clean_uri = uri.strip().strip("/")
                    tested_uris.add(uri)

                    # Try to match it with a known route
                    matched = route_lookup.get(uri) or route_lookup.get("/" + uri) or route_lookup.get(uri + "/") or route_lookup.get("/" + uri + "/")
                    if matched:
                        route_name = matched.get("Route Name")
                        cid = matched.get("Component ID")
                        if route_name:
                            tested_routes.add(route_name)
                        elif cid:
                            tested_routes.add(cid)

                component_id = generate_component_id("TST")

                entry = {
                    "Component ID": component_id,
                    "Test Class": class_name,
                    "Namespace": namespace,
                    "Defined At": rel_path,
                    "Test Methods": ", ".join(test_methods),
                    "Tested URIs": ", ".join(sorted(tested_uris)),
                    "Tested Routes": ", ".join(sorted(tested_routes)),
                }

                results.append(entry)

            except Exception as e:
                err_entry = {
                    "Component ID": "ERR",
                    "Test Class": file,
                    "Namespace": "",
                    "Defined At": rel_path,
                    "Test Methods": "",
                    "Tested URIs": "",
                    "Tested Routes": "",
                    "Error": str(e)
                }
                results.append(err_entry)
                logging.error(f" Error scanning test file {file}: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Tests: {len(results)} entries with enhanced coverage")

    return "Tests", results

def scan_validation_rules():
    logging.info(" Scanning for Validation Rules ...")
    results = []

    # 1. Scan FormRequest classes in app/Http/Requests
    req_dir = PROJECT_ROOT / "app" / "Http" / "Requests"
    if req_dir.exists():
        for root, _, files in os.walk(req_dir):
            for file in files:
                if not file.endswith(".php"):
                    continue
                path = Path(root) / file
                rel_path = str(path.relative_to(PROJECT_ROOT))

                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()

                class_name = find_php_class_name(content)
                rules = re.findall(r"['\"](\w+)['\"]\s*=>\s*['\"](.+?)['\"]", content)
                if rules:
                    for field, rule in rules:
                        results.append({
                            "Component ID": generate_component_id("VAL"),
                            "Class": class_name,
                            "Type": "FormRequest",
                            "Field": field,
                            "Rule": rule,
                            "Defined At": rel_path
                        })

    # 2. Scan inline validations in Controllers (validate([...]))
    for ctrl in sheet_data.get("Controllers", []):
        try:
            ctrl_path = PROJECT_ROOT / ctrl["Defined At"]
            if not ctrl_path.exists():
                continue

            with open(ctrl_path, "r", encoding="utf-8") as f:
                content = f.read()

            validations = re.findall(r'request\(\)->validate\(\[([^\]]+)\]', content, re.DOTALL)
            for val_block in validations:
                rule_lines = re.findall(r"['\"](\w+)['\"]\s*=>\s*['\"](.+?)['\"]", val_block)
                for field, rule in rule_lines:
                    results.append({
                        "Component ID": generate_component_id("VAL"),
                        "Class": ctrl["Controller"],
                        "Type": "Inline",
                        "Field": field,
                        "Rule": rule,
                        "Defined At": ctrl["Defined At"]
                    })

        except Exception as e:
            logging.warning(f"Validation scan error in {ctrl.get('Controller')}: {e}")

    return "Validation Rules", results


# ========== 🛡️ SECURITY AUDIT SCANNER ==========

def scan_security():
    logging.info(" Running Security Audit ...")
    results = []
    env_file = PROJECT_ROOT / ".env"

    if env_file.exists():
        try:
            with open(env_file, "r", encoding="utf-8") as f:
                lines = f.readlines()

            secrets = []
            for line in lines:
                if any(keyword in line.lower() for keyword in ["secret", "key", "password", "token", "access"]):
                    secrets.append(line.strip())

            if secrets:
                entry = {
                    "Component ID": generate_component_id("SEC"),
                    ".env Keys": ", ".join(secrets[:10]),
                    "File": ".env",
                }
                results.append(entry)
        except Exception as e:
            logging.error(f" Failed to read .env: {e}")

    # Scan config/app.php
    config_file = PROJECT_ROOT / "config" / "app.php"
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read()

            risky_lines = []
            if "APP_KEY" in content or "base64:" in content:
                risky_lines.append("Contains base64 encoded APP_KEY")

            if risky_lines:
                entry = {
                    "Component ID": generate_component_id("SEC"),
                    ".env Keys": ", ".join(risky_lines),
                    "File": "config/app.php",
                }
                results.append(entry)
        except Exception as e:
            logging.error(f" Failed to scan config/app.php: {e}")

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Security audit entries: {len(results)} issues")

    return "Security Audit", results


# ========== 📂 FOLDER STRUCTURE SCANNER ==========

def scan_folder_structure():
    logging.info(" Scanning folder structure ...")
    results = []

    for root, dirs, files in os.walk(PROJECT_ROOT):
        rel_root = Path(root).relative_to(PROJECT_ROOT)
        folder_name = str(rel_root) if rel_root != Path('.') else "."

        try:
            num_files = len(files)
            total_size = sum(os.path.getsize(Path(root) / f) for f in files if os.path.isfile(Path(root) / f))

            entry = {
                "Component ID": generate_component_id("DIR"),
                "Folder": folder_name,
                "Files": num_files,
                "Size": readable_file_size(total_size),
            }

            results.append(entry)

        except Exception as e:
            logging.error(f" Error scanning folder {folder_name}: {e}")
            continue

    if DEBUG_PRINT_MINIMAL:
        print(f"✓ Scanned Folder Structure: {len(results)} folders")

    return "Folder Structure", results

# ========== 📦 DEPENDENCY SCANNER ==========

def scan_dependencies():
    logging.info(" Scanning composer.json for dependencies ...")
    composer_file = PROJECT_ROOT / "composer.json"
    results = []

    if not composer_file.exists():
        logging.warning(" composer.json not found.")
        return "Dependencies", []

    try:
        with open(composer_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        dependencies = data.get("require", {})
        dev_dependencies = data.get("require-dev", {})

        for pkg, version in dependencies.items():
            results.append({
                "Component ID": generate_component_id("DEP"),
                "Package": pkg,
                "Version": version,
                "Type": "Runtime"
            })

        for pkg, version in dev_dependencies.items():
            results.append({
                "Component ID": generate_component_id("DEP"),
                "Package": pkg,
                "Version": version,
                "Type": "Dev"
            })

        if DEBUG_PRINT_MINIMAL:
            print(f"✓ Scanned Dependencies: {len(results)} packages")

    except Exception as e:
        logging.error(f" Error reading composer.json: {e}")

    return "Dependencies", results

def scan_validation_rules():
    logging.info(" Scanning Validation Rules ...")
    results = []

    # --- Scan FormRequest files ---
    request_dir = PROJECT_ROOT / "app" / "Http" / "Requests"
    if request_dir.exists():
        for root, _, files in os.walk(request_dir):
            for file in files:
                if not file.endswith(".php"):
                    continue
                path = Path(root) / file
                rel_path = str(path.relative_to(PROJECT_ROOT))
                try:
                    content = path.read_text(encoding="utf-8")
                    class_name = find_php_class_name(content)
                    rules = re.findall(r"['\"](\w+)['\"]\s*=>\s*['\"](.+?)['\"]", content)
                    for field, rule in rules:
                        results.append({
                            "Component ID": generate_component_id("VAL"),
                            "Class": class_name,
                            "Type": "FormRequest",
                            "Field": field,
                            "Rule": rule,
                            "Defined At": rel_path
                        })
                except Exception as e:
                    logging.warning(f"Validation scan error in {file}: {e}")

    # --- Scan Controllers for inline validate([...]) ---
    for ctrl in sheet_data.get("Controllers", []):
        ctrl_path = PROJECT_ROOT / ctrl["Defined At"]
        if not ctrl_path.exists():
            continue
        try:
            content = ctrl_path.read_text(encoding="utf-8")
            validations = re.findall(r'request\(\)->validate\(\[([^\]]+)\]', content, re.DOTALL)
            for val_block in validations:
                rule_lines = re.findall(r"['\"](\w+)['\"]\s*=>\s*['\"](.+?)['\"]", val_block)
                for field, rule in rule_lines:
                    results.append({
                        "Component ID": generate_component_id("VAL"),
                        "Class": ctrl["Controller"],
                        "Type": "Inline",
                        "Field": field,
                        "Rule": rule,
                        "Defined At": ctrl["Defined At"]
                    })
        except Exception as e:
            logging.warning(f"Inline validation scan error in {ctrl['Controller']}: {e}")

    return "Validation Rules", results

def scan_config_usage():
    logging.info(" Scanning Config Usage ...")
    results = []
    app_dir = PROJECT_ROOT / "app"

    for root, _, files in os.walk(app_dir):
        for file in files:
            if not file.endswith(".php"):
                continue
            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                content = path.read_text(encoding="utf-8")
                matches = re.findall(r"config\(['\"](.+?)['\"]\)", content)
                for match in matches:
                    results.append({
                        "Component ID": generate_component_id("CFG"),
                        "Config Key": match,
                        "Used In": rel_path
                    })
            except Exception as e:
                logging.warning(f"Error reading {file}: {e}")

    return "Config Usage", results

def scan_authorization_usage():
    logging.info(" Scanning Authorization Usage ...")
    results = []

    # Scan Controllers for Gate or authorize()
    for ctrl in sheet_data.get("Controllers", []):
        path = PROJECT_ROOT / ctrl["Defined At"]
        if not path.exists():
            continue
        try:
            content = path.read_text(encoding="utf-8")
            lines = content.splitlines()

            for i, line in enumerate(lines):
                # authorize('ability', ...)
                auth_matches = re.findall(r'authorize\([\'"](.+?)[\'"]', line)
                for ability in auth_matches:
                    results.append({
                        "Source Type": "Controller",
                        "Component": ctrl["Controller"],
                        "File": ctrl["Defined At"],
                        "Line": i + 1,
                        "Method/Directive": "authorize()",
                        "Target Ability": ability,
                        "Context": extract_context(line)
                    })

                # Gate::allows('...', ...)
                gate_allows = re.findall(r'Gate::allows\([\'"](.+?)[\'"]', line)
                for ability in gate_allows:
                    results.append({
                        "Source Type": "Controller",
                        "Component": ctrl["Controller"],
                        "File": ctrl["Defined At"],
                        "Line": i + 1,
                        "Method/Directive": "Gate::allows()",
                        "Target Ability": ability,
                        "Context": extract_context(line)
                    })

                # Gate::denies
                gate_denies = re.findall(r'Gate::denies\([\'"](.+?)[\'"]', line)
                for ability in gate_denies:
                    results.append({
                        "Source Type": "Controller",
                        "Component": ctrl["Controller"],
                        "File": ctrl["Defined At"],
                        "Line": i + 1,
                        "Method/Directive": "Gate::denies()",
                        "Target Ability": ability,
                        "Context": extract_context(line)
                    })

        except Exception as e:
            logging.warning(f"Failed to scan controller {ctrl['Controller']} for authorization: {e}")

    # Scan Views for @can / @cannot / @canany
    for view in sheet_data.get("Views", []):
        path = PROJECT_ROOT / view["File"]
        if not path.exists():
            continue
        try:
            content = path.read_text(encoding="utf-8")
            lines = content.splitlines()
            for i, line in enumerate(lines):
                for directive in ["@can", "@cannot", "@canany"]:
                    matches = re.findall(fr"{directive}\([\'\"](.+?)[\'\"]", line)
                    for ability in matches:
                        results.append({
                            "Source Type": "Blade View",
                            "Component": view["View"],
                            "File": view["File"],
                            "Line": i + 1,
                            "Method/Directive": directive,
                            "Target Ability": ability,
                            "Context": extract_context(line)
                        })
        except Exception as e:
            logging.warning(f"Failed to scan view {view['View']} for authorization: {e}")

    return "Authorization Map", results

def extract_context(line):
    ctx = ""
    ctx_match = re.findall(r",\s*(\$[\w\-\.>]+)", line)
    if ctx_match:
        ctx = ctx_match[0]
    return ctx

def scan_data_exporters():
    logging.info(" Scanning Export Usage (PDF/Excel)...")
    results = []

    for root, _, files in os.walk(PROJECT_ROOT / "app"):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                content = path.read_text(encoding="utf-8")
                lines = content.splitlines()
                for i, line in enumerate(lines):

                    # PDF: PDF::loadView('view.name')
                    pdf_matches = re.findall(r'PDF::loadView\([\'"](.+?)[\'"]', line)
                    for view in pdf_matches:
                        results.append({
                            "Export Type": "PDF",
                            "Export Class": "",
                            "View Used": view,
                            "Used In File": rel_path,
                            "Line": i + 1,
                            "Method": "PDF::loadView"
                        })

                    # Excel: Excel::download(new SomeExportClass)
                    excel_matches = re.findall(r'Excel::download\(\s*new\s+(\w+)', line)
                    for export_class in excel_matches:
                        results.append({
                            "Export Type": "Excel",
                            "Export Class": export_class,
                            "View Used": "",
                            "Used In File": rel_path,
                            "Line": i + 1,
                            "Method": "Excel::download"
                        })

                    # Excel::store
                    store_matches = re.findall(r'Excel::store\(\s*new\s+(\w+)', line)
                    for export_class in store_matches:
                        results.append({
                            "Export Type": "Excel",
                            "Export Class": export_class,
                            "View Used": "",
                            "Used In File": rel_path,
                            "Line": i + 1,
                            "Method": "Excel::store"
                        })

            except Exception as e:
                logging.warning(f"Export scan failed in {file}: {e}")

    return "Data Exporters", results

def scan_model_relationships():
    logging.info(" Scanning Model Relationships ...")
    results = []

    model_dir = PROJECT_ROOT / "app" / "Models"
    if not model_dir.exists():
        logging.warning("Model directory not found.")
        return "Model Relationships", []

    for root, _, files in os.walk(model_dir):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                content = path.read_text(encoding="utf-8")
                class_name = find_php_class_name(content)
                methods = re.findall(r'public function (\w+)\(\)\s*\{([^}]+)\}', content, re.DOTALL)

                for method_name, body in methods:
                    match = re.search(r'(hasOne|hasMany|belongsTo|belongsToMany|morphTo|morphMany)\s*\(\s*(?:[\'"]([\w\\]+)[\'"])?', body)
                    if match:
                        rel_type = match.group(1)
                        related_model = match.group(2) if match.group(2) else ""
                        results.append({
                            "Model": class_name,
                            "Relationship Type": rel_type,
                            "Related Model": related_model,
                            "Method Name": method_name,
                            "Defined In File": rel_path
                        })

            except Exception as e:
                logging.warning(f"Error reading model {file}: {e}")

    return "Model Relationships", results

def scan_seeders_and_factories():
    logging.info(" Scanning Seeders and Factories ...")
    results = []

    # FACTORIES
    factory_dir = PROJECT_ROOT / "database" / "factories"
    if factory_dir.exists():
        for root, _, files in os.walk(factory_dir):
            for file in files:
                if not file.endswith(".php"):
                    continue
                path = Path(root) / file
                rel_path = str(path.relative_to(PROJECT_ROOT))

                try:
                    content = path.read_text(encoding="utf-8")
                    class_name = find_php_class_name(content)
                    model_match = re.search(r"->define\(\s*([A-Za-z0-9_:\\]+)::class", content) or \
                                  re.search(r"factory\s*\(\s*([A-Za-z0-9_:\\]+)::class", content)
                    model = model_match.group(1) if model_match else ""
                    results.append({
                        "Type": "Factory",
                        "Class Name": class_name,
                        "Used Model": model,
                        "Called In Seeder": "",
                        "Defined At": rel_path
                    })
                except Exception as e:
                    logging.warning(f"Error in factory {file}: {e}")

    # SEEDERS
    seeder_dir = PROJECT_ROOT / "database" / "seeders"
    if seeder_dir.exists():
        for root, _, files in os.walk(seeder_dir):
            for file in files:
                if not file.endswith(".php"):
                    continue
                path = Path(root) / file
                rel_path = str(path.relative_to(PROJECT_ROOT))

                try:
                    content = path.read_text(encoding="utf-8")
                    class_name = find_php_class_name(content)

                    # Seeder calling another seeder
                    called_seeders = re.findall(r'\$this->call\(\s*([A-Za-z0-9_\\]+)::class', content)

                    # Factory calls like: User::factory()->create()
                    factory_calls = re.findall(r'([A-Za-z0-9_\\]+)::factory\(\)', content)

                    if called_seeders:
                        for callee in called_seeders:
                            results.append({
                                "Type": "Seeder",
                                "Class Name": class_name,
                                "Used Model": "",
                                "Called In Seeder": callee,
                                "Defined At": rel_path
                            })

                    if factory_calls:
                        for model in factory_calls:
                            results.append({
                                "Type": "Seeder",
                                "Class Name": class_name,
                                "Used Model": model,
                                "Called In Seeder": "",
                                "Defined At": rel_path
                            })

                    if not called_seeders and not factory_calls:
                        results.append({
                            "Type": "Seeder",
                            "Class Name": class_name,
                            "Used Model": "",
                            "Called In Seeder": "",
                            "Defined At": rel_path
                        })

                except Exception as e:
                    logging.warning(f"Error in seeder {file}: {e}")

    return "Seeders & Factories", results

def scan_livewire_components():
    logging.info(" Scanning Livewire Components ...")
    results = []

    livewire_dir = PROJECT_ROOT / "app" / "Http" / "Livewire"
    if livewire_dir.exists():
        for root, _, files in os.walk(livewire_dir):
            for file in files:
                if not file.endswith(".php"):
                    continue
                path = Path(root) / file
                rel_path = str(path.relative_to(PROJECT_ROOT))
                try:
                    content = path.read_text(encoding="utf-8")
                    class_name = find_php_class_name(content)
                    lines = content.splitlines()

                    for i, line in enumerate(lines):
                        if "emit(" in line or "emitTo(" in line:
                            match = re.search(r"emit(To)?\(\s*[\'\"](.+?)[\'\"]", line)
                            if match:
                                results.append({
                                    "Type": "Emit",
                                    "Component/Class": class_name,
                                    "Tag/Usage": match.group(2),
                                    "Found In": "Livewire Component",
                                    "File": rel_path,
                                    "Line": i + 1
                                })
                        if "render(" in line or "mount(" in line:
                            results.append({
                                "Type": "Method",
                                "Component/Class": class_name,
                                "Tag/Usage": "render/mount",
                                "Found In": "Livewire Component",
                                "File": rel_path,
                                "Line": i + 1
                            })

                except Exception as e:
                    logging.warning(f"Error scanning Livewire class {file}: {e}")

    # Scan Blade Views for @livewire or <livewire:... />
    for view in sheet_data.get("Views", []):
        path = PROJECT_ROOT / view["File"]
        if not path.exists():
            continue
        try:
            lines = path.read_text(encoding="utf-8").splitlines()
            for i, line in enumerate(lines):
                # @livewire('component-name')
                matches1 = re.findall(r"@livewire\(['\"](.+?)['\"]", line)
                for match in matches1:
                    results.append({
                        "Type": "Directive",
                        "Component/Class": match,
                        "Tag/Usage": "@livewire",
                        "Found In": "Blade View",
                        "File": view["File"],
                        "Line": i + 1
                    })

                # <livewire:component-name />
                matches2 = re.findall(r"<livewire:(.+?)[\s/>]", line)
                for match in matches2:
                    results.append({
                        "Type": "Component",
                        "Component/Class": match,
                        "Tag/Usage": f"<livewire:{match}>",
                        "Found In": "Blade View",
                        "File": view["File"],
                        "Line": i + 1
                    })

        except Exception as e:
            logging.warning(f"Livewire view parse failed: {e}")

    return "Livewire Components", results

def scan_service_dependencies():
    logging.info(" Scanning Dependency Injection & Service Container usage ...")
    results = []

    for root, _, files in os.walk(PROJECT_ROOT / "app"):
        for file in files:
            if not file.endswith(".php"):
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))
            try:
                content = path.read_text(encoding="utf-8")
                class_name = find_php_class_name(content)
                lines = content.splitlines()

                for i, line in enumerate(lines):
                    # Constructor DI
                    if "__construct" in line:
                        param_matches = re.findall(r'([A-Za-z0-9_\\]+)\s+\$[a-zA-Z_][\w]*', line)
                        for param in param_matches:
                            if "\\" in param and not param.startswith(("Illuminate", "App\\Models")):
                                results.append({
                                    "Class": class_name,
                                    "Injected Class": param,
                                    "Type": "Constructor DI",
                                    "File": rel_path,
                                    "Line": i + 1
                                })

                    # app()->make(SomeService::class)
                    make_matches = re.findall(r'app\(\)->make\(([^)]+)\)', line)
                    for match in make_matches:
                        cleaned = match.replace("::class", "").strip(" '\"")
                        results.append({
                            "Class": class_name,
                            "Injected Class": cleaned,
                            "Type": "app()->make()",
                            "File": rel_path,
                            "Line": i + 1
                        })

                    # resolve(Foo::class)
                    resolve_matches = re.findall(r'resolve\(([^)]+)\)', line)
                    for match in resolve_matches:
                        cleaned = match.replace("::class", "").strip(" '\"")
                        results.append({
                            "Class": class_name,
                            "Injected Class": cleaned,
                            "Type": "resolve()",
                            "File": rel_path,
                            "Line": i + 1
                        })

            except Exception as e:
                logging.warning(f"DI scan failed in {file}: {e}")

    return "Service Dependencies", results

def scan_env_usage():
    logging.info(" Scanning env('...') usage ...")
    results = []
    skipped_non_utf8_files = []

    for root, _, files in os.walk(PROJECT_ROOT):
        for file in files:
            if not file.endswith(".php"):
                continue

            # 🛑 Skip known problematic files (binary/legacy)
            if file.startswith("from.") and file.endswith(".php"):
                skipped_non_utf8_files.append(file)
                continue

            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                lines = path.read_text(encoding="utf-8").splitlines()
                for i, line in enumerate(lines):
                    matches = re.findall(r"env\(['\"](.+?)['\"]", line)
                    for key in matches:
                        results.append({
                            "Key": key,
                            "Used In File": rel_path,
                            "Line": i + 1,
                            "Context Snippet": line.strip()
                        })
            except UnicodeDecodeError:
                skipped_non_utf8_files.append(file)
            except Exception as e:
                logging.warning(f"env() scan failed in {file}: {e}")

    if skipped_non_utf8_files:
        logging.warning(f"Skipped {len(skipped_non_utf8_files)} non-UTF8 or incompatible files (e.g., {skipped_non_utf8_files[:3]}...)")

    return "Environment Usage", results

def scan_events_and_listeners():
    logging.info(" Scanning Events and Listeners ...")
    results = []

    for root, _, files in os.walk(PROJECT_ROOT / "app"):
        for file in files:
            if not file.endswith(".php"):
                continue
            path = Path(root) / file
            rel_path = str(path.relative_to(PROJECT_ROOT))

            try:
                content = path.read_text(encoding="utf-8")
                lines = content.splitlines()

                for i, line in enumerate(lines):
                    # event(new SomeEvent)
                    match = re.findall(r"event\s*\(\s*new\s+([A-Za-z0-9_\\]+)", line)
                    for evt in match:
                        results.append({
                            "Type": "Event",
                            "Event/Listener Class": evt,
                            "Action": "event(new ...)",
                            "Triggered In File": rel_path,
                            "Line": i + 1,
                            "Description": "Manual dispatch"
                        })

                    # broadcast(new SomeEvent)
                    match2 = re.findall(r"broadcast\s*\(\s*new\s+([A-Za-z0-9_\\]+)", line)
                    for evt in match2:
                        results.append({
                            "Type": "Event",
                            "Event/Listener Class": evt,
                            "Action": "broadcast(...)",
                            "Triggered In File": rel_path,
                            "Line": i + 1,
                            "Description": "Broadcast event"
                        })

                    # handle() in Listener
                    if "function handle(" in line and "Listener" in rel_path:
                        class_name = find_php_class_name(content)
                        results.append({
                            "Type": "Listener",
                            "Event/Listener Class": class_name,
                            "Action": "handle()",
                            "Triggered In File": rel_path,
                            "Line": i + 1,
                            "Description": "Listener function"
                        })

            except Exception as e:
                logging.warning(f"Failed scanning event/listener in {file}: {e}")

    return "Events & Listeners", results

def parallel_scan(scan_funcs):
    results = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_func = {executor.submit(func): name for name, func in scan_funcs}
        for future in as_completed(future_to_func):
            try:
                sheet_name, sheet_rows = future.result()
                sheet_data[sheet_name] = sheet_rows
                results.append(sheet_name)
                logging.info(f" Completed {sheet_name} ({len(sheet_rows)} rows)")
            except Exception as e:
                logging.warning(f"[✗] Scan failed: {future_to_func[future]} - {e}")
    return results


# ========== 📑 TABLE OF CONTENTS GENERATOR ==========

def generate_toc():
    logging.info(" Generating TOC ...")
    toc = []

    for idx, sheet in enumerate(SHEET_ORDER, start=1):
        row_count = len(sheet_data.get(sheet, []))
        toc.append({
            "Sheet #": idx,
            "Sheet Name": sheet,
            "Entries": row_count,
            "Status": "✅" if row_count > 0 else "❌"
        })

    return "TOC", toc

def generate_manual_qa_sheet():
    """Manual QA/DevOps checklist for untrackable Laravel scenarios"""
    logging.info("Generating Manual QA Checklist ...")

    checklist = [
        {"Category": "Runtime", "Checklist Item": "Check logs for unexpected exceptions", "Status": "☐"},
        {"Category": "Runtime", "Checklist Item": "Test controller logic with edge-case inputs", "Status": "☐"},
        {"Category": "Runtime", "Checklist Item": "Ensure Sentry/Bugsnag is reporting exceptions", "Status": "☐"},

        {"Category": "External API", "Checklist Item": "Verify all external APIs are reachable and valid", "Status": "☐"},
        {"Category": "External API", "Checklist Item": "Ensure timeout/fallback logic exists for API calls", "Status": "☐"},

        {"Category": "CI/CD", "Checklist Item": "CI/CD pipelines pass on all branches", "Status": "☐"},
        {"Category": "CI/CD", "Checklist Item": "Deployment environment matches `.env.example`", "Status": "☐"},

        {"Category": "Environment", "Checklist Item": "Check Laravel version compatibility in `composer.lock`", "Status": "☐"},
        {"Category": "Environment", "Checklist Item": "No hardcoded secrets in configs or code", "Status": "☐"},

        {"Category": "UI/Browser", "Checklist Item": "Test UI in Chrome, Firefox, Safari", "Status": "☐"},
        {"Category": "UI/Browser", "Checklist Item": "Verify responsive design on mobile", "Status": "☐"},
        {"Category": "UI/Browser", "Checklist Item": "Inspect for missing assets/CSS in browser dev tools", "Status": "☐"},
    ]

    return "Manual QA", checklist


# ========== 📊 DASHBOARD GENERATOR ==========

def generate_dashboard():
    logging.info(" Generating Dashboard ...")
    dashboard = []

    total_components = sum(len(v) for v in sheet_data.values())
    controller_count = len(sheet_data.get("Controllers", []))
    model_count = len(sheet_data.get("Models", []))
    view_count = len(sheet_data.get("Views", []))
    route_count = len(sheet_data.get("Web Routes", [])) + len(sheet_data.get("API Routes", []))

    dashboard.append({"Metric": "Total Components", "Value": total_components})
    dashboard.append({"Metric": "Controllers", "Value": controller_count})
    dashboard.append({"Metric": "Models", "Value": model_count})
    dashboard.append({"Metric": "Views", "Value": view_count})
    dashboard.append({"Metric": "Routes (Web + API)", "Value": route_count})
    dashboard.append({"Metric": "Scanned At", "Value": get_current_timestamp()})
    dashboard.append({"Metric": "Project Root", "Value": str(PROJECT_ROOT)})

    return "Dashboard", dashboard


# ========== 🧠 MASTER REFERENCE SHEET ==========

def generate_master_reference():
    logging.info(" Compiling Master Reference ...")
    master = []

    for cid, data in component_index.items():
        master.append({
            "Component ID": cid,
            "Sheet": data.get("sheet", ""),
            "Row": data.get("row", ""),
            "Name": data.get("name", ""),
            "File": data.get("file", "")
        })

    master.sort(key=lambda x: x["Component ID"])
    return "Master Reference", master

# ========== 🔗 HYPERLINK + VALIDATION ENGINE (continued) ==========

from openpyxl.utils import quote_sheetname

def hyperlink_cell(ws, row_idx, col_idx, target_sheet, target_row, label="→"):
    """Safely create a clickable cell that links to a target cell in another sheet"""
    cell = ws.cell(row=row_idx, column=col_idx, value=label)
    cell.font = Font(color="0563C1", underline="single")
    cell.hyperlink = f"#{quote_sheetname(target_sheet)}!A{target_row}"
    return cell

def build_component_index(wb):
    """Build comprehensive component index with validation"""
    global component_index
    component_index.clear()

    logging.info("Building component index with validation...")

    # ✅ Type mapping for better naming (plural-safe)
    TYPE_MAP = {
        "Web Routes": "Route",
        "API Routes": "Route",
        "Controllers": "Controller",
        "Models": "Model",
        "Views": "View",
        "JavaScript": "Script",
        "Middleware": "Middleware",
        "Migrations": "Migration",
        "Database Schema": "Table",
        "Service Providers": "ServiceProvider",
        "Config Files": "Config",
        "Tests": "Test",
        "Security Audit": "Security",
        "Folder Structure": "Folder",
        "Dependencies": "Package",
        "Master Reference": "Reference",
        "Route References": "RouteLink",
        "Component Relationships": "Link"
    }

    # Initialize lookups structure
    lookups = {
        'routes': {
            'by_uri': {},
            'by_name': {},
            'by_controller': defaultdict(list)
        },
        'controllers': {
            'by_name': {},
            'by_full_class': {},
            'by_namespace': defaultdict(list)
        },
        'models': {
            'by_name': {},
            'by_full_class': {},
            'by_table': {}
        },
        'views': {
            'by_name': {},
            'by_path': {}
        },
        # ✅ Generic lookups
        'view': defaultdict(list),
        'file': defaultdict(list),
        'model': defaultdict(list)
    }

    # Track stats
    stats = {
        'components_added': 0,
        'sheets_processed': 0,
        'rows_skipped': 0,
        'invalid_entries': 0,
        'duplicates': 0
    }

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_ORDER:
            logging.debug(f"Skipping unmapped sheet: {sheet_name}")
            continue

        try:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                logging.debug(f"Skipping empty sheet: {sheet_name}")
                continue

            headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

            # Detect component ID column
            id_col = None
            for id_field in ['Component ID', 'ID', 'component_id']:
                if id_field in headers:
                    id_col = headers.index(id_field) + 1
                    break

            if not id_col:
                logging.warning(f"No ID column found in {sheet_name}")
                continue

            stats['sheets_processed'] += 1

            for row in ws.iter_rows(min_row=2):
                try:
                    component_id = str(ws.cell(row=row[0].row, column=id_col).value or "").strip()

                    if not component_id or component_id == "ERR":
                        stats['rows_skipped'] += 1
                        continue

                    if component_id in component_index:
                        logging.warning(f"Duplicate Component ID '{component_id}' found in {sheet_name} row {row[0].row}")
                        stats['duplicates'] += 1

                    component_data = {
                        'sheet': sheet_name,
                        'row': row[0].row,
                        'name': '',
                        'file': '',
                        'type': TYPE_MAP.get(sheet_name, sheet_name),
                        'id': component_id
                    }

                    # Route-specific parsing
                    if sheet_name in ['Web Routes', 'API Routes']:
                        process_route_row(ws, row, headers, component_data, lookups)
                    elif sheet_name == 'Controllers':
                        process_controller_row(ws, row, headers, component_data, lookups)
                    elif sheet_name == 'Models':
                        process_model_row(ws, row, headers, component_data, lookups)
                    elif sheet_name == 'Views':
                        process_view_row(ws, row, headers, component_data, lookups)
                    else:
                        logging.debug(f"No custom parser for {sheet_name}; storing minimal info")

                    # Add to index (overwrite-safe)
                    component_index[component_id] = component_data
                    stats['components_added'] += 1

                except Exception as e:
                    stats['invalid_entries'] += 1
                    logging.warning(f"Error processing row {row[0].row} in {sheet_name}: {str(e)}")
                    continue

        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {str(e)}")
            continue

    # ✅ Attach metadata — mark _lookups as internal
    component_index["_lookups"] = lookups   # <--- ⚠️ skip this in iteration
    component_index["_stats"] = stats

    logging.info(f"Index built with {stats['components_added']} components")
    logging.info(f"Sheets processed: {stats['sheets_processed']}")
    logging.info(f"Rows skipped: {stats['rows_skipped']}")
    logging.info(f"Invalid entries: {stats['invalid_entries']}")
    logging.info(f"Duplicate IDs: {stats['duplicates']}")

    return True

def process_route_row(ws, row, headers, component_data, lookups):
    """Process a row from routes sheet"""
    if 'URI' in headers:
        uri = ws.cell(row=row[0].row, column=headers.index('URI') + 1).value or ""
        component_data['name'] = uri
        if uri:
            lookups['routes']['by_uri'][uri] = component_data['id']
            
    if 'Route Name' in headers:
        route_name = ws.cell(row=row[0].row, column=headers.index('Route Name') + 1).value or ""
        if route_name:
            lookups['routes']['by_name'][route_name] = component_data['id']
            
    if 'Controller' in headers:
        ctrl = ws.cell(row=row[0].row, column=headers.index('Controller') + 1).value or ""
        component_data['file'] = ctrl
        if ctrl:
            ctrl_name = ctrl.split('@')[0]
            lookups['routes']['by_controller'][ctrl_name].append(component_data['id'])

def process_controller_row(ws, row, headers, component_data, lookups):
    """Process a row from controllers sheet"""
    if 'Controller' in headers:
        ctrl_name = ws.cell(row=row[0].row, column=headers.index('Controller') + 1).value or ""
        component_data['name'] = ctrl_name
        if ctrl_name:
            lookups['controllers']['by_name'][ctrl_name] = component_data['id']
            
    if 'Full Class' in headers:
        full_class = ws.cell(row=row[0].row, column=headers.index('Full Class') + 1).value or ""
        if full_class:
            lookups['controllers']['by_full_class'][full_class] = component_data['id']
            if '\\' in full_class:
                namespace = full_class.split('\\')[0]
                lookups['controllers']['by_namespace'][namespace].append(component_data['id'])

def process_view_row(ws, row, headers, component_data, lookups):
    """Process a row from the Views sheet and populate all lookup dictionaries"""
    view_name = ""
    file_path = ""

    if 'View' in headers:
        view_name = str(ws.cell(row=row[0].row, column=headers.index('View') + 1).value or "").strip()
        component_data['name'] = view_name
        if view_name:
            lookups['views']['by_name'][view_name] = component_data['id']
            lookups['view'][view_name].append(component_data['id'])
            lookups['view'][view_name.replace('.', '/') + '.blade.php'].append(component_data['id'])
            lookups['view'][view_name.lower()].append(component_data['id'])

    if 'File' in headers:
        file_path = str(ws.cell(row=row[0].row, column=headers.index('File') + 1).value or "").strip()
        component_data['file'] = file_path
        if file_path:
            lookups['views']['by_path'][file_path] = component_data['id']
            lookups['file'][file_path].append(component_data['id'])
            lookups['file'][file_path.lower()].append(component_data['id'])

def process_model_row(ws, row, headers, component_data, lookups):
    """Process a row from the Models sheet and update model lookups"""
    if 'Model' in headers:
        model_name = str(ws.cell(row=row[0].row, column=headers.index('Model') + 1).value or "").strip()
        component_data['name'] = model_name
        if model_name:
            lookups['models']['by_name'][model_name] = component_data['id']
            lookups['model'][model_name].append(component_data['id'])
            lookups['model'][model_name.lower()].append(component_data['id'])

    if 'Full Class' in headers:
        full_class = str(ws.cell(row=row[0].row, column=headers.index('Full Class') + 1).value or "").strip()
        if full_class:
            lookups['models']['by_full_class'][full_class] = component_data['id']

    if 'Defines Table' in headers:
        table_name = str(ws.cell(row=row[0].row, column=headers.index('Defines Table') + 1).value or "").strip()
        if table_name:
            lookups['models']['by_table'][table_name] = component_data['id']

def verify_component_index():
    """Debug function to check component index integrity"""
    logging.info("Verifying component index...")
    total = len(component_index)
    if total == 0:
        logging.error("COMPONENT INDEX IS EMPTY!")
        return False
    
    # Count components per sheet
    from collections import defaultdict
    sheet_counts = defaultdict(int)
    for cid, data in component_index.items():
        sheet_counts[data['sheet']] += 1
    
    logging.info(f"Index contains {total} components across {len(sheet_counts)} sheets")
    for sheet, count in sheet_counts.items():
        logging.info(f"- {sheet}: {count} components")
    
    # Check some sample entries
    sample = list(component_index.items())[:3]
    for cid, data in sample:
        logging.info(f"Sample: {cid} -> {data['sheet']} (row {data['row']})")
    
    return True

def verify_component_index_integrity(wb):
    """Verify the component index matches actual worksheet data"""
    logging.info("Verifying component index integrity...")
    
    issues_found = 0
    valid_sheets = set(wb.sheetnames)
    
    # Check for components pointing to non-existent sheets
    for cid, data in list(component_index.items()):
        # Skip metadata entries and ensure data has 'sheet' key
        if cid.startswith('_'):
            continue
            
        if not isinstance(data, dict) or 'sheet' not in data:
            logging.warning(f"Invalid component entry: {cid}")
            issues_found += 1
            continue
            
        if data['sheet'] not in valid_sheets:
            logging.warning(f"Component {cid} points to missing sheet: {data['sheet']}")
            issues_found += 1
    
    # Check for orphaned components (in index but not in sheets)
    sheet_components = defaultdict(set)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1] if cell.value]
        
        if "Component ID" in headers:
            id_col = headers.index("Component ID") + 1
            for row in ws.iter_rows(min_row=2):
                cid = ws.cell(row=row[0].row, column=id_col).value
                if cid:
                    sheet_components[sheet_name].add(cid)
    
    for cid, data in list(component_index.items()):
        if cid.startswith('_'):
            continue
            
        if not isinstance(data, dict) or 'sheet' not in data:
            continue
            
        if data['sheet'] in sheet_components and cid not in sheet_components[data['sheet']]:
            logging.warning(f"Orphaned component {cid} not found in sheet {data['sheet']}")
            issues_found += 1
    
    if issues_found == 0:
        logging.info("Component index verification passed")
    else:
        logging.error(f"Found {issues_found} component index issues")
    
    return issues_found == 0

def debug_component_index(wb):
    """Debug helper to verify component index and sheet names"""
    if not component_index:
        logging.warning("Component index is empty!")
        return

    # Verify first 5 components
    sample_components = list(component_index.items())[:5]
    logging.info("Sample component index entries:")
    for cid, data in sample_components:
        logging.info(f"  {cid} -> Sheet: '{data['sheet']}', Row: {data['row']}")

    # Verify sheet names match exactly
    sheet_names = set(wb.sheetnames)
    index_sheets = set(data['sheet'] for data in component_index.values())
    missing_sheets = index_sheets - sheet_names
    if missing_sheets:
        logging.warning(f"Missing sheets in workbook: {missing_sheets}")

def create_single_hyperlink(ws, cell, target_id, source_sheet, stats):
    """Create one hyperlink with comprehensive validation"""
    if not target_id or target_id.startswith('_'):
        stats['invalid_refs'] += 1
        return False

    target = component_index.get(target_id)
    if not isinstance(target, dict) or 'sheet' not in target or 'row' not in target:
        stats['errors'] += 1
        stats['missing_targets'][target_id] += 1
        return False

    # Skip same-sheet links
    if target['sheet'] == source_sheet:
        stats['same_sheet_skips'] += 1
        return False

    # Verify target sheet exists
    if target['sheet'] == source_sheet:
        stats['same_sheet_skips'] += 1
        logging.debug(f" Skipped same-sheet link: {source_sheet} → {target_id} on row {target['row']}")
        return False

    try:
        safe_sheet = target['sheet'].replace("'", "''")
        hyperlink = f"#'{safe_sheet}'!A{target['row']}"
        cell.hyperlink = hyperlink
        cell.font = Font(color="0563C1", underline="single")

        comment_text = f"Links to: {target['sheet']}\n{target.get('name', target.get('file', target_id))}"
        if not cell.comment:
            cell.comment = Comment(comment_text, "Laravel Tracker")

        return True
    except Exception as e:
        logging.debug(f"Failed to create hyperlink: {str(e)}")
        stats['errors'] += 1
        return False

def log_hyperlink_stats(stats):
    """Generate detailed hyperlinking report with method breakdown and ref context"""
    logging.info("\n Hyperlinking Results:")
    logging.info(f" Successfully created: {stats['success']}")
    logging.info(f" Errors encountered: {stats['errors']}")

    if stats['missing_targets']:
        logging.warning(f"\n Missing Targets ({len(stats['missing_targets'])} unique):")
        for target, count in sorted(stats['missing_targets'].items(), key=lambda x: x[1], reverse=True)[:10]:
            logging.warning(f"- {target} (referenced {count}x)")

    if stats['missing_sheets']:
        logging.warning(f"\n Missing Sheets ({len(stats['missing_sheets'])}):")
        for sheet in sorted(stats['missing_sheets']):
            logging.warning(f"- {sheet}")

    if stats['resolution_method']:
        logging.info("\n Resolution Method Breakdown:")
        total = sum(stats['resolution_method'].values())
        for method, count in stats['resolution_method'].items():
            pct = (count / total) * 100 if total else 0
            logging.info(f" - {method}: {count} ({pct:.1f}%)")

    # 🔽 Write unresolved refs with cleanup context
    try:
        with open("unresolved_refs.txt", "w", encoding="utf-8") as f:
            f.write("Unresolved References (with cleaned + normalized forms):\n\n")
            for ref, count in sorted(stats['missing_targets'].items(), key=lambda x: x[1], reverse=True):
                cleaned = clean_blade_placeholders(ref).strip()
                norm = cleaned.replace('\\', '/').replace('.blade.php', '').replace('/', '.').strip()
                f.write(f"{ref} → cleaned: {cleaned}, normalized: {norm} (x{count})\n")
        logging.info(" Detailed unresolved references written to unresolved_refs.txt")
    except Exception as e:
        logging.error(f"Failed to write unresolved_refs.txt: {e}")


def suggest_solutions(stats):
    """Provide actionable solutions based on error patterns"""
    if not stats['errors']:
        return
    
    logging.info("\nSuggested Solutions:")
    
    if stats['missing_targets']:
        logging.info("1. For missing component references:")
        logging.info("   - Verify the components exist in the scanned data")
        logging.info("   - Check naming conventions match between references and components")
        
        # Detect common pattern in missing targets
        sample_missing = next(iter(stats['missing_targets']))
        if '{' in sample_missing:
            logging.info("   - Template routes detected - ensure base routes exist")
    
    if stats['missing_sheets']:
        logging.info("2. For missing sheets:")
        logging.info("   - Add these sheets to SHEET_ORDER list:")
        for sheet in sorted(stats['missing_sheets']):
            logging.info(f"     - '{sheet}'")
    
    if stats['invalid_refs']:
        logging.info("3. For invalid references:")
        logging.info("   - Check for empty or 'ERR' values in relationship columns")

    if stats['errors'] > 20:
        logging.info("4. Consider disabling HIGHLIGHT_EMPTY_CELLS or reviewing relationships manually.")    

def write_unresolved_refs(stats, filename="unresolved_refs.txt"):
    """Output unresolved references to file"""
    try:
        with open(filename, "w", encoding="utf-8") as f:
            f.write("🔍 Unresolved References\n")
            f.write("=========================\n\n")
            for ref, count in sorted(stats['missing_targets'].items(), key=lambda x: x[1], reverse=True):
                f.write(f"{ref} (referenced {count}x)\n")
        logging.info(f" Unresolved references written to: {filename}")
    except Exception as e:
        logging.error(f"Failed to write unresolved refs: {e}")

def verify_hyperlinks(wb):
    """Verify all hyperlinks point to valid locations"""
    if not hasattr(wb, 'sheetnames'):  # Check if wb is a valid workbook
        logging.error("Invalid workbook provided for verification")
        return False

    valid_sheets = set(wb.sheetnames)
    broken_links = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                    try:
                        # Extract target from hyperlink
                        target = cell.hyperlink.target.split("!")[0].strip("#'")
                        if target not in valid_sheets:
                            broken_links += 1
                            logging.warning(f"Broken hyperlink in {sheet_name}!{cell.coordinate} -> {target}")
                    except Exception as e:
                        broken_links += 1
                        logging.debug(f"Invalid hyperlink format in {sheet_name}!{cell.coordinate}: {e}")
    
    status = broken_links == 0
    logging.info(f"Hyperlink verification {'passed' if status else 'failed'} with {broken_links} broken links")
    return status

def check_missing_references():
    """Debug why references are missing"""
    from collections import defaultdict
    missing = defaultdict(int)
    
    for sheet in SHEET_ORDER:
        if sheet not in sheet_data:
            continue
            
        for row in sheet_data[sheet]:
            for field in ["Controller ID", "Model ID", "View ID"]:
                if field in row and row[field]:
                    ref_id = str(row[field]).strip()
                    if ref_id not in component_index:
                        missing[field] += 1
    
    if missing:
        logging.error("Missing reference analysis:")
        for field, count in missing.items():
            logging.error(f"- {count} missing {field} references")
        
        # Show sample missing IDs
        samples = []
        for sheet in SHEET_ORDER[:3]:  # Check first 3 sheets
            if sheet in sheet_data:
                for row in sheet_data[sheet][:5]:  # First 5 rows
                    if "Controller ID" in row and row["Controller ID"]:
                        cid = row["Controller ID"]
                        if cid not in component_index:
                            samples.append(cid)
                            if len(samples) >= 3:
                                break
        if samples:
            logging.error(f"Sample missing IDs: {', '.join(samples)}")                       

# ========== 🧾 EXCEL WRITER ==========

def write_sheet(wb, sheet_name, rows):
    if not rows:
        return

    # Truncate sheet name to Excel's 31-character limit
    ws = wb.create_sheet(title=sheet_name[:31])
    headers = list(rows[0].keys())

    # Add Link column if Component ID exists
    if "Component ID" in headers and "🔗 Link" not in headers:
        headers.append("🔗 Link")
        for row in rows:
            row["🔗 Link"] = row["Component ID"]

    # Write headers with standard formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = get_header_fill()
        cell.font = get_bold_font()
        cell.alignment = get_center_alignment()
        cell.border = get_border()

    # Special formatting for Route References sheet
    if sheet_name == "Route References":
        # Highlight important columns
        highlight_cols = {
            "Route Name": "FFF2CC",  # Light orange
            "Referenced In": "E2EFDA",  # Light green
            "File": "DEEBF7"  # Light blue
        }
        
        # Apply special formatting to data rows
        for row_idx, data_row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, 1):
                val = data_row.get(header, "")
                cell = write_cell(ws, row_idx, col_idx, val)
                
                # Apply background colors to specific columns
                if header in highlight_cols:
                    cell.fill = PatternFill(
                        start_color=highlight_cols[header],
                        end_color=highlight_cols[header],
                        fill_type="solid"
                    )
                
                # Make route names bold
                if header == "Route Name" and val:
                    cell.font = get_bold_font()
                
                cell.border = get_border()
        
        # Add auto-filter for easy filtering
        ws.auto_filter.ref = ws.dimensions
        
    else:
        # Standard data writing for other sheets
        for row_idx, data_row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, 1):
                val = data_row.get(header, "")
                cell = write_cell(ws, row_idx, col_idx, val)
                cell.border = get_border()

    # Apply conditional formatting for error cells
    if "Error" in headers:
        error_col = headers.index("Error") + 1
        red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=error_col, max_col=error_col):
            for cell in row:
                if cell.value:
                    cell.fill = red_fill

    apply_auto_column_width(ws, headers)
    freeze_header(ws)

    # Special formatting for Views sheet
    if sheet_name == "Views":
        # Highlight views with route references
        if "Route References" in headers:
            ref_col = headers.index("Route References") + 1
            yellow_fill = PatternFill(start_color="FFFFCC", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ref_col, max_col=ref_col):
                for cell in row:
                    if cell.value:
                        cell.fill = yellow_fill
                        cell.font = Font(bold=True)

    # Optional: Special formatting for Manual QA checklist
    if sheet_name == "Manual QA":
        if "Status" in headers:
            status_col = headers.index("Status") + 1
            gray_fill = PatternFill(start_color="EAEAEA", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col, max_col=status_col):
                for cell in row:
                    cell.fill = gray_fill
                    cell.alignment = get_center_alignment()
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    

def create_hyperlinks(wb):
    """Create cross-sheet hyperlinks with robust error handling and reporting"""
    logging.info("Creating hyperlinks with enhanced validation...")

    if not component_index or "_lookups" not in component_index:
        logging.error("Component index not properly built!")
        return {
            'success': 0,
            'errors': 1,
            'missing_targets': defaultdict(int),
            'same_sheet_skips': 0,
            'invalid_refs': 0,
            'missing_sheets': set(),
            'resolution_method': defaultdict(int)
        }

    stats = {
        'success': 0,
        'errors': 0,
        'missing_targets': defaultdict(int),
        'same_sheet_skips': 0,
        'invalid_refs': 0,
        'missing_sheets': set(),
        'resolution_method': defaultdict(int)
    }

    success_log = []

    RELATIONSHIP_COLUMNS = [
        "Component ID", "🔗 Link", "Controller ID",
        "Used By Routes", "Uses Models", "Uses Views",
        "Used In Controllers", "Route References",
        "Related To", "Dependencies"
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        relationship_cols = [
            (idx + 1, header)
            for idx, header in enumerate(headers)
            if any(keyword in header for keyword in RELATIONSHIP_COLUMNS)
        ]

        for col_idx, header in relationship_cols:
            for row in ws.iter_rows(min_row=2):
                cell = ws.cell(row=row[0].row, column=col_idx)
                if not cell.value:
                    continue

                refs = parse_cell_references(cell.value)
                for ref in refs:
                    raw = str(ref).strip()

                    # Skip garbage or Blade placeholders
                    if not raw or raw in {'ERR', '', '"', "'"} or '{{' in raw or '}}' in raw:
                        stats['invalid_refs'] += 1
                        continue

                    # Clean and normalize
                    cleaned_ref = clean_blade_placeholders(raw).strip()
                    normalized_ref = cleaned_ref.replace('\\', '/').replace('.blade.php', '').replace('/', '.').strip()

                    # Try resolving target
                    target_id = find_target_id(cleaned_ref, component_index["_lookups"])
                    method_used = "cleaned"

                    if not target_id:
                        target_id = find_target_id(normalized_ref, component_index["_lookups"])
                        method_used = "normalized"

                    if not target_id:
                        stats['missing_targets'][raw] += 1
                        stats['errors'] += 1
                        logging.debug(f"No match for: {raw} → cleaned='{cleaned_ref}', normalized='{normalized_ref}'")
                        continue

                    stats['resolution_method'][method_used] += 1

                    if create_single_hyperlink(ws, cell, target_id, sheet_name, stats):
                        stats['success'] += 1
                        target = component_index.get(target_id, {})
                        success_log.append(
                            f"{sheet_name}!{cell.coordinate} → {target.get('sheet', '?')}!A{target.get('row', '?')} (component_id: {target_id})"
                        )

    # Write audit log of created hyperlinks
    if success_log:
        try:
            with open("hyperlink_map.txt", "w", encoding="utf-8") as f:
                f.write("Successful Cross-Sheet Hyperlinks:\n\n")
                for entry in success_log:
                    f.write(f"{entry}\n")
            logging.info(f"Hyperlink audit log written to hyperlink_map.txt ({len(success_log)} links)")
        except Exception as e:
            logging.error(f"Failed to write hyperlink_map.txt: {e}")

    # Log summary and suggestions
    log_hyperlink_stats(stats)
    suggest_solutions(stats)
    write_unresolved_refs(stats)
    log_resolution_breakdown(stats)

    return stats

def parse_cell_references(value):
    """Parse multiple references from a cell with flexible delimiters"""
    if not value:
        return []

    if isinstance(value, str):
        delimiters = [',', '|', '\n', ';']
        for delim in delimiters:
            if delim in value:
                return [r.strip() for r in value.split(delim) if r.strip()]
        return [value.strip()]
    
    return [str(value)]  

def build_route_lookups():
    """Build lookup dictionaries for routes"""
    lookups = {
        'routes_by_uri': {},
        'routes_by_name': {}
    }
    
    for sheet_name in ['Web Routes', 'API Routes']:
        for route in sheet_data.get(sheet_name, []):
            if 'URI' in route:
                # Normalize URI by removing leading/trailing slashes
                normalized_uri = route['URI'].strip('/')
                
                # Store multiple URI formats
                uri_formats = {
                    normalized_uri,
                    f"/{normalized_uri}",
                    f"/{normalized_uri}/",
                    f"{normalized_uri}/"
                }
                
                for uri in uri_formats:
                    lookups['routes_by_uri'][uri] = route['Component ID']
                
            if 'Route Name' in route and route['Route Name']:
                lookups['routes_by_name'][route['Route Name']] = route['Component ID']
    
    return lookups

def build_controller_lookups():
    """Build lookup dictionaries for controllers"""
    return {
        'controllers_by_name': {
            controller['Controller']: controller['Component ID']
            for controller in sheet_data.get('Controllers', [])
        }
    }

def is_relationship_column(header):
    """Check if column contains relationship data"""
    relationship_keywords = [
        'ID', 'Link', 'Routes', 'Models', 
        'Views', 'Controllers', 'References'
    ]
    return any(kw in str(header) for kw in relationship_keywords)

def find_target_id(ref, lookups):
    """Find target ID using multiple matching strategies with enhanced fallback logic"""

    if not ref or not isinstance(ref, str):
        logging.debug(f"Invalid reference: {ref}")
        return None

    # --- Step 1: Skip garbage/static files ---
    garbage = {'', 'ERR', '"', "'"}
    if ref.strip() in garbage or ref.lower().endswith(('.css', '.js', '.png', '.jpg', '.jpeg')):
        logging.debug(f"Skipped invalid/static reference: {ref}")
        return None

    # --- Step 2: Known alias remapping ---
    ALIAS_MAP = {
        "plexpaypasswordchange": "user.password.change",
        "expensedownload": "expense.index",
        "billdeskfinalrecieptwithouttax": "invoice.receipt",
        "generatetax-pdfsunmi": "invoice.tax.sunmi",
        "salesreportfinal": "report.sales",
        "search": "search.index",
        "download": "files.download",
        "gethistory": "user.history",
        "daily_report": "report.daily"
    }

    if hasattr(find_target_id, "_alias_seen") is False:
        seen = {}
        duplicates = {}
        for k, v in ALIAS_MAP.items():
            if k in seen:
                duplicates[k] = (seen[k], v)
            seen[k] = v
        if duplicates:
            logging.warning("Duplicate keys found in ALIAS_MAP:")
            for key, vals in duplicates.items():
                logging.warning(f" - {key}: {vals[0]} overwritten by {vals[1]}")
        find_target_id._alias_seen = True

    ref = ALIAS_MAP.get(ref.strip(), ref.strip())
    original_ref = ref.strip()

    # --- Step 3: Normalize Blade refs ---
    ref_clean = (
        re.sub(r'\{\{.*?\}\}', '', original_ref)
        .replace('\\', '/')
        .replace('.blade.php', '')
        .replace('/', '.')
        .strip('.')
    ).lower()

    # --- Step 4: Direct Component ID match ---
    if original_ref in component_index:
        return original_ref
    if ref_clean in component_index:
        return ref_clean

    # --- Step 5: Route URI matching (cleaned + variants) ---
    base_uri = re.sub(r'\{.*?\}', '', ref).strip('/').lower()

    uri_variants = {
        ref_clean,
        base_uri,
        f"/{base_uri}",
        f"{base_uri}/",
        f"api/{base_uri}",
        f"api/{base_uri}/",
        base_uri.replace('/', '.'),
        base_uri.replace('.', '/')
    }

    for uri in uri_variants:
        if uri in lookups['routes']['by_uri']:
            return lookups['routes']['by_uri'][uri]

    # --- Step 6: Route name match ---
    if ref in lookups['routes']['by_name']:
        return lookups['routes']['by_name'][ref]
    if '.' in ref:
        base = ref.split('.')[0]
        if base in lookups['routes']['by_name']:
            return lookups['routes']['by_name'][base]

    # --- Step 7: Controller match ---
    if ref in lookups['controllers']['by_name']:
        return lookups['controllers']['by_name'][ref]
    if '\\' in ref:
        short = ref.split('\\')[-1]
        if short in lookups['controllers']['by_name']:
            return lookups['controllers']['by_name'][short]
    if ref.lower().endswith('controller'):
        base = ref[:-10]
        if base in lookups['controllers']['by_name']:
            return lookups['controllers']['by_name'][base]

    # --- Step 8: View match (dot, slash, normalized) ---
    view_variants = {
        ref,
        ref.replace('/', '.'),
        ref.replace('.', '/'),
        ref.replace('.', '/') + '.blade.php',
        ref_clean
    }
    for view_key in view_variants:
        if view_key in lookups['view']:
            return lookups['view'][view_key][0]

    # --- Step 9: File match ---
    if ref in lookups['file']:
        return lookups['file'][ref][0]
    if ref_clean in lookups['file']:
        return lookups['file'][ref_clean][0]

    # --- Step 10: Model match ---
    if ref in lookups['model']:
        return lookups['model'][ref][0]
    if ref_clean in lookups['model']:
        return lookups['model'][ref_clean][0]

    # --- Step 11: Fuzzy fallback ---
    for category in ['view', 'file', 'model']:
        for key in lookups.get(category, {}):
            if key and (ref_clean.startswith(key) or key in ref_clean):
                logging.debug(f"Fuzzy match in '{category}': {ref} -> {key}")
                return lookups[category][key][0]

    for uri in lookups.get('routes', {}).get('by_uri', {}):
        if uri and (ref_clean.startswith(uri) or uri in ref_clean):
            logging.debug(f"Fuzzy route match: {ref} -> {uri}")
            return lookups['routes']['by_uri'][uri]

    logging.debug(f"No match found for reference: {original_ref}")
    return None

def track_missing_reference(ref, stats, route_lookups):
    """Track missing references with categorization"""
    if any(c in ref for c in ['/', '\\', '{']):  # Likely route URI (including template vars)
        stats['missing_routes'][ref] += 1
    elif ref[0].isupper():  # Likely controller/model
        stats['missing_controllers'][ref] += 1
    else:
        stats['errors'] += 1

def create_hyperlink(ws, cell, target_id, source_sheet):
    """Create a single hyperlink with validation"""
    target = component_index[target_id]
    
    # Skip same-sheet links
    if target['sheet'] == source_sheet:
        return False
        
    # Check target exists
    if target['sheet'] not in ws.parent.sheetnames:
        return False

    try:
        safe_sheet = target['sheet'].replace("'", "''")
        cell.hyperlink = f"#'{safe_sheet}'!A{target['row']}"
        cell.font = Font(color="0563C1", underline="single")
        return True
    except Exception:
        return False

def log_resolution_breakdown(stats):
    if 'resolution_method' not in stats:
        return

    logging.info("\n Resolution Method Breakdown:")
    total = sum(stats['resolution_method'].values())
    for method, count in sorted(stats['resolution_method'].items(), key=lambda x: x[1], reverse=True):
        pct = (count / total * 100) if total else 0
        logging.info(f"  {method:<12} -> {count} refs ({pct:.1f}%)")

def suggest_route_fix(uri, route_lookups):
    """Provide suggestions for missing route fixes"""
    suggestions = []
    normalized = uri.strip('/')
    
    # Handle template routes like "daily_report/{{ $uid }}"
    if '{' in uri:
        base_route = uri.split('{')[0].rstrip('/')
        if base_route in route_lookups['routes_by_uri']:
            suggestions.append(f"Template route matches base: {base_route}")
        else:
            # Try removing all template segments
            clean_route = re.sub(r'\{[^}]+\}', '', uri).replace('//', '/').strip('/')
            if clean_route in route_lookups['routes_by_uri']:
                suggestions.append(f"Try route without templates: {clean_route}")
    
    # Check for common prefix mismatches
    for prefix in ['api/', 'admin/', 'app/', '']:
        test_route = f"{prefix}{normalized}"
        if test_route in route_lookups['routes_by_uri']:
            suggestions.append(f"Try with prefix '{prefix}'")
            break
    
    # Check for trailing/leading slash variants
    for variant in [f"/{normalized}", f"/{normalized}/", f"{normalized}/"]:
        if variant in route_lookups['routes_by_uri']:
            suggestions.append(f"Exists as: {variant}")
            break
    
    if suggestions:
        logging.info("  Suggestions: " + ", ".join(suggestions))
    else:
        logging.info("  No similar routes found in index")

def normalize_class_name(name):
    if not name:
        return ""
    return name.split("\\")[-1].replace(";", "").strip()

# ========== 🔄 RUN ALL SCANNERS ==========
def run_all_scanners():
    scanners = [
        scan_web_routes,
        scan_api_routes,
        scan_controllers,
        scan_models,
        scan_views,
        scan_javascript,
        scan_middleware,
        scan_migrations,
        scan_schema,
        scan_config,
        scan_service_providers,
        scan_tests,
        scan_security,
        scan_folder_structure,
        scan_dependencies,
        scan_validation_rules,     
        scan_config_usage,
        scan_data_exporters,
        scan_seeders_and_factories,
        scan_livewire_components,
        scan_service_dependencies,
        scan_env_usage,
    ]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(scanner): scanner.__name__ for scanner in scanners}
        for future in as_completed(futures):
            try:
                sheet, rows = future.result()
                sheet_data[sheet] = rows
            except Exception as e:
                logging.error(f" Scanner failed: {futures[future]} — {e}")


def build_component_relationships():
    """Build complete two-way relationships between components with enhanced matching"""
    logging.info("Building comprehensive component relationships...")
    
    # Initialize relationship maps with tracking for stats
    relationships = {
        'controller_routes': defaultdict(set),
        'controller_models': defaultdict(set),
        'controller_views': defaultdict(set),
        'model_controllers': defaultdict(set),
        'view_controllers': defaultdict(set),
        'route_controllers': defaultdict(set),
        'js_routes': defaultdict(set),
        'view_components': defaultdict(set)
    }

    # Initialize comprehensive lookup dictionaries with route lookups first
    lookups = {
        'routes': {
            'by_uri': {},
            'by_name': {},
            'by_controller': defaultdict(list)
        },
        'controllers': {
            'by_name': {},
            'by_full_class': {},
            'by_namespace': defaultdict(list)
        },
        'models': {
            'by_name': {},
            'by_full_class': {},
            'by_table': {}
        },
        'views': {
            'by_name': {},
            'by_path': {}
        }
    }

    # First build route lookups to ensure they're available for other relationships
    try:
        for route_sheet in ["Web Routes", "API Routes"]:
            if route_sheet not in sheet_data:
                continue
                
            for route in sheet_data[route_sheet]:
                if not isinstance(route, dict):
                    continue
                
                # Add URI lookup
                if 'URI' in route and route['URI']:
                    normalized_uri = route['URI'].strip('/')
                    lookups['routes']['by_uri'][normalized_uri] = route['Component ID']
                    # Add variants with/without leading/trailing slashes
                    lookups['routes']['by_uri'][f"/{normalized_uri}"] = route['Component ID']
                    lookups['routes']['by_uri'][f"/{normalized_uri}/"] = route['Component ID']
                    lookups['routes']['by_uri'][f"{normalized_uri}/"] = route['Component ID']
                
                # Add route name lookup
                if 'Route Name' in route and route['Route Name']:
                    lookups['routes']['by_name'][route['Route Name']] = route['Component ID']
                
                # Add controller-route mapping
                if 'Controller' in route and route['Controller']:
                    ctrl_name = route['Controller'].split('@')[0]
                    lookups['routes']['by_controller'][ctrl_name].append(route['Component ID'])
    except Exception as e:
        logging.error(f"Failed to build route lookups: {str(e)}")
        return False

    # Populate controller lookups with error handling
    try:
        for controller in sheet_data.get("Controllers", []):
            if not isinstance(controller, dict):
                continue
                
            cid = controller.get("Component ID")
            name = controller.get("Controller")
            full_class = controller.get("Full Class", "")
            
            if not cid or not name:
                continue
                
            lookups['controllers']['by_name'][name] = cid
            if full_class:
                lookups['controllers']['by_full_class'][full_class] = cid
                # Store with and without namespace
                if '\\' in full_class:
                    namespace = full_class.split('\\')[0]
                    base_name = full_class.split('\\')[-1]
                    lookups['controllers']['by_namespace'][namespace].append(cid)
                    lookups['controllers']['by_name'][base_name] = cid
    except Exception as e:
        logging.error(f"Failed to build controller lookups: {str(e)}")
        return False

    # Populate model lookups with error handling
    try:
        for model in sheet_data.get("Models", []):
            if not isinstance(model, dict):
                continue
                
            mid = model.get("Component ID")
            name = model.get("Model")
            full_class = model.get("Full Class", "")
            table = model.get("Defines Table", "")
            
            if not mid or not name:
                continue
                
            lookups['models']['by_name'][name] = mid
            if full_class:
                lookups['models']['by_full_class'][full_class] = mid
                if '\\' in full_class:
                    lookups['models']['by_name'][full_class.split('\\')[-1]] = mid
            if table:
                lookups['models']['by_table'][table] = mid
    except Exception as e:
        logging.error(f"Failed to build model lookups: {str(e)}")
        return False

    # Populate view lookups with error handling
    try:
        for view in sheet_data.get("Views", []):
            if not isinstance(view, dict):
                continue
                
            vid = view.get("Component ID")
            name = view.get("View")
            file_path = view.get("File", "")
            
            if not vid or not name:
                continue
                
            # Store with dot notation (view.name)
            lookups['views']['by_name'][name] = vid
            
            # Also store with path notation (view/name.blade.php)
            path_name = name.replace('.', '/') + '.blade.php'
            lookups['views']['by_name'][path_name] = vid
            
            # Store full file path if available
            if file_path:
                lookups['views']['by_path'][file_path] = vid
    except Exception as e:
        logging.error(f"Failed to build view lookups: {str(e)}")
        return False

    # Build relationships with error handling
    try:
        # 1. Enhanced Route ↔ Controller linking
        for route_sheet in ["Web Routes", "API Routes"]:
            for route in sheet_data.get(route_sheet, []):
                if not isinstance(route, dict):
                    continue
                    
                if not route.get("Controller"):
                    continue
                    
                ctrl_ref = route["Controller"].split('@')[0]
                cid = find_matching_component(ctrl_ref, lookups['controllers'])
                
                if cid:
                    route["Controller ID"] = cid
                    relationships['controller_routes'][cid].add(route["Component ID"])
                    relationships['route_controllers'][route["Component ID"]].add(cid)

        # 2. Enhanced Controller ↔ Model linking
        for controller in sheet_data.get("Controllers", []):
            cid = controller["Component ID"]
            
            # Check both "Uses Models" and model type hints in code
            model_refs = set()
            if "Uses Models" in controller and controller["Uses Models"]:
                model_refs.update(m.strip() for m in controller["Uses Models"].split(",") if m.strip())
            
            # Additional model detection from method signatures
            if "Methods" in controller:  # If we've parsed methods
                for method in controller["Methods"].split(','):
                    if 'Model' in method:  # Simplified detection
                        model_refs.add(method.split('Model')[1].split()[0])
            
            for model_ref in model_refs:
                mid = find_matching_component(model_ref, lookups['models'])
                if mid:
                    relationships['controller_models'][cid].add(mid)
                    relationships['model_controllers'][mid].add(cid)

        # 3. Enhanced Controller ↔ View linking
        for controller in sheet_data.get("Controllers", []):
            cid = controller["Component ID"]
            
            # Check both explicit view references and implicit ones
            view_refs = set()
            if "Uses Views" in controller and controller["Uses Views"]:
                view_refs.update(v.strip() for v in controller["Uses Views"].split(",") if v.strip())
            
            # Additional view detection from controller code
            if "View References" in controller:  # If we've parsed view calls
                view_refs.update(controller["View References"].split(','))
            
            for view_ref in view_refs:
                vid = find_matching_component(view_ref, lookups['views'])
                if vid:
                    relationships['controller_views'][cid].add(vid)
                    relationships['view_controllers'][vid].add(cid)

        # 4. JavaScript ↔ Route linking
        for js in sheet_data.get("JavaScript", []):
            if "Fetch/Route URIs" not in js:
                continue
                
            for uri in js["Fetch/Route URIs"].split(','):
                uri = uri.strip()
                if uri in lookups['routes']['by_uri']:
                    relationships['js_routes'][js["Component ID"]].add(lookups['routes']['by_uri'][uri])

        # 5. View Component relationships
        for view in sheet_data.get("Views", []):
            if "Includes" not in view:
                continue
                
            for component in view["Includes"].split(','):
                component = component.strip()
                if component in lookups['views']['by_name']:
                    relationships['view_components'][view["Component ID"]].add(
                        lookups['views']['by_name'][component]
                    )

        # Update all relationship fields
        update_relationship_fields(relationships)

        # Store lookups and relationships for later use
        component_index["_lookups"] = lookups
        component_index["_relationships"] = relationships

        # Log comprehensive stats
        log_relationship_stats(relationships)
        
        return True

    except Exception as e:
        logging.error(f"Failed to build relationships: {str(e)}")
        return False

def find_matching_component(ref, lookup_dict):
    """Find matching component using multiple matching strategies"""
    # 1. Direct name match
    if ref in lookup_dict['by_name']:
        return lookup_dict['by_name'][ref]
    
    # 2. Full class match
    if 'by_full_class' in lookup_dict and ref in lookup_dict['by_full_class']:
        return lookup_dict['by_full_class'][ref]
    
    # 3. Try with normalized names
    normalized = ref.split('\\')[-1]  # Remove namespace
    if normalized in lookup_dict['by_name']:
        return lookup_dict['by_name'][normalized]
    
    # 4. Try removing common suffixes
    for suffix in ['Controller', 'Model', 'Service', 'Repository']:
        if ref.endswith(suffix):
            base_name = ref[:-len(suffix)]
            if base_name in lookup_dict['by_name']:
                return lookup_dict['by_name'][base_name]
    
    return None

def update_relationship_fields(relationships):
    """Update all relationship fields in sheet data"""
    # Update Controllers with their relationships
    for controller in sheet_data.get("Controllers", []):
        cid = controller["Component ID"]
        
        # Routes
        if cid in relationships['controller_routes']:
            route_ids = relationships['controller_routes'][cid]
            controller["Used By Routes"] = ", ".join(sorted(route_ids))
        
        # Models
        if cid in relationships['controller_models']:
            model_ids = relationships['controller_models'][cid]
            model_names = []
            for mid in sorted(model_ids):
                model = next((m for m in sheet_data.get("Models", []) if m["Component ID"] == mid), None)
                if model:
                    model_names.append(model["Model"])
            controller["Uses Models"] = ", ".join(model_names)
        
        # Views
        if cid in relationships['controller_views']:
            view_ids = relationships['controller_views'][cid]
            view_names = []
            for vid in sorted(view_ids):
                view = next((v for v in sheet_data.get("Views", []) if v["Component ID"] == vid), None)
                if view:
                    view_names.append(view["View"])
            controller["Uses Views"] = ", ".join(view_names)

    # Update Models with their controllers
    for model in sheet_data.get("Models", []):
        mid = model["Component ID"]
        if mid in relationships['model_controllers']:
            ctrl_ids = relationships['model_controllers'][mid]
            ctrl_names = []
            for cid in sorted(ctrl_ids):
                ctrl = next((c for c in sheet_data.get("Controllers", []) if c["Component ID"] == cid), None)
                if ctrl:
                    ctrl_names.append(ctrl["Controller"])
            model["Used In Controllers"] = ", ".join(ctrl_names)

    # Update Views with their controllers
    for view in sheet_data.get("Views", []):
        vid = view["Component ID"]
        if vid in relationships['view_controllers']:
            ctrl_ids = relationships['view_controllers'][vid]
            ctrl_names = []
            for cid in sorted(ctrl_ids):
                ctrl = next((c for c in sheet_data.get("Controllers", []) if c["Component ID"] == cid), None)
                if ctrl:
                    ctrl_names.append(ctrl["Controller"])
            view["Used By Controllers"] = ", ".join(ctrl_names)

def log_relationship_stats(relationships):
    """Log detailed relationship statistics"""
    stats = {
        'Controllers with routes': len(relationships['controller_routes']),
        'Controllers with models': len(relationships['controller_models']),
        'Controllers with views': len(relationships['controller_views']),
        'Models with controllers': len(relationships['model_controllers']),
        'Views with controllers': len(relationships['view_controllers']),
        'JavaScript files with routes': len(relationships['js_routes']),
        'Views with components': len(relationships['view_components'])
    }
    
    logging.info("Component Relationship Statistics:")
    for category, count in stats.items():
        logging.info(f"- {category}: {count}")
    
    # Show sample relationships
    if relationships['controller_routes']:
        sample_cid = next(iter(relationships['controller_routes']))
        sample_ctrl = next((c for c in sheet_data.get("Controllers", []) 
                          if c["Component ID"] == sample_cid), None)
        if sample_ctrl:
            logging.info(f"\nSample Controller Relationships:")
            logging.info(f"Controller: {sample_ctrl['Controller']}")
            if sample_cid in relationships['controller_routes']:
                logging.info(f"Routes: {len(relationships['controller_routes'][sample_cid])}")
            if sample_cid in relationships['controller_models']:
                logging.info(f"Models: {len(relationships['controller_models'][sample_cid])}")
            if sample_cid in relationships['controller_views']:
                logging.info(f"Views: {len(relationships['controller_views'][sample_cid])}")

def get_component_name(component_id):
    """Helper to get component name from ID"""
    for sheet in sheet_data.values():
        for item in sheet:
            if item.get("Component ID") == component_id:
                return item.get("Controller") or item.get("Model") or item.get("View") or item.get("File") or component_id
    return component_id

def normalize_class_name(name):
    """Normalize class names for comparison"""
    if not name:
        return ""
    return name.split('\\')[-1].replace(";", "").strip()

def generate_relationship_report():
    """Generate a report of all component relationships"""
    relationships = []
    
    # Controller -> Model relationships
    for controller in sheet_data.get("Controllers", []):
        if "Uses Models" in controller and controller["Uses Models"]:
            relationships.append({
                "Source Type": "Controller",
                "Source": controller["Controller"],
                "Relationship": "Uses Model",
                "Target Type": "Model",
                "Target": controller["Uses Models"]
            })
    
    # Controller -> View relationships
    for controller in sheet_data.get("Controllers", []):
        if "Uses Views" in controller and controller["Uses Views"]:
            relationships.append({
                "Source Type": "Controller",
                "Source": controller["Controller"],
                "Relationship": "Uses View",
                "Target Type": "View",
                "Target": controller["Uses Views"]
            })
    
    # Route -> Controller relationships
    for route_sheet in ["Web Routes", "API Routes"]:
        for route in sheet_data.get(route_sheet, []):
            if "Controller" in route and route["Controller"]:
                relationships.append({
                    "Source Type": "Route",
                    "Source": route["URI"],
                    "Relationship": "Calls",
                    "Target Type": "Controller",
                    "Target": route["Controller"]
                })
    
    # JavaScript -> Route relationships
    for js_file in sheet_data.get("JavaScript", []):
        if "Uses URI" in js_file and js_file["Uses URI"]:
            relationships.append({
                "Source Type": "JavaScript",
                "Source": js_file["File"],
                "Relationship": "Calls",
                "Target Type": "Route",
                "Target": js_file["Uses URI"]
            })
    
    return "Component Relationships", relationships

def stable_component_id(prefix, path):
    digest = hashlib.md5(str(path).encode()).hexdigest()[:8]
    return f"{prefix}-{digest}"

# ========== 🚀 MAIN FUNCTION ==========
def main():
    print("📁 Laravel Tracker Started...")
    try:
        # 1. INITIALIZATION
        load_config()
        
        # 2. SCAN PROJECT COMPONENTS
        print("🔍 Scanning project components...")
        run_all_scanners()
        
        # 3. BUILD RELATIONSHIPS
        print("🧩 Building component relationships...")
        if not build_component_relationships():
            logging.error("Failed to build component relationships")
            return False
        
        # 4. GENERATE METADATA SHEETS
        print("📊 Generating metadata sheets...")
        # Step 1: Scans that must run synchronously (routes, migrations, etc.)
        sheet_data.update({
            "Web Routes": scan_web_routes()[1],
            "API Routes": scan_api_routes()[1],
            "Migrations": scan_migrations()[1],
            "Middleware": scan_middleware()[1],
            "Config Files": scan_config()[1],
            "Service Providers": scan_service_providers()[1],
            "Dependencies": scan_dependencies()[1],
        })

        # Step 2: Parallel scan of expensive modules
        parallel_scan([
            ("Controllers", scan_controllers),
            ("Models", scan_models),
            ("Views", scan_views),
            ("JavaScript", scan_javascript),
            ("Tests", scan_tests),
            ("Validation Rules", scan_validation_rules),
            ("Livewire Components", scan_livewire_components),
        ])

        # Step 3: Dependent scans (must run after controllers/views/routes are populated)
        sheet_data.update({
            "Route References": generate_route_reference_report()[1],
            "Blade Hierarchy": generate_blade_hierarchy()[1],
            "Authorization Map": scan_authorization_usage()[1],
            "Model Relationships": scan_model_relationships()[1],
            "Events & Listeners": scan_events_and_listeners()[1],
            "Service Dependencies": scan_service_dependencies()[1],
            "Environment Usage": scan_env_usage()[1],
            "Config Usage": scan_config_usage()[1],
            "Security Audit": scan_security()[1],
            "Database Schema": scan_schema()[1],
            "Folder Structure": scan_folder_structure()[1],
            "Seeders & Factories": scan_seeders_and_factories()[1],
            "Data Exporters": scan_data_exporters()[1],
            "Component Relationships": generate_relationship_report()[1],
        })

        # Step 4: Final sheets
        sheet_data.update({
            "Master Reference": generate_master_reference()[1],
            "Dashboard": generate_dashboard()[1],
            "TOC": generate_toc()[1],
            "Manual QA": generate_manual_qa_sheet()[1],
        })

        
        # 5. CREATE AND PREPARE WORKBOOK
        print("📑 Creating workbook...")
        wb = Workbook()
        del wb["Sheet"]  # Remove default sheet
        
        # 6. WRITE ALL SHEETS TO WORKBOOK
        print("✍️ Writing sheets to workbook...")
        for sheet in SHEET_ORDER:
            if sheet in sheet_data:
                write_sheet(wb, sheet, sheet_data[sheet])
        
        # 7. BUILD COMPONENT INDEX
        print("📇 Building component index...")
        if not build_component_index(wb):
            logging.error("Failed to build component index")
            return False

        # 8. VERIFY COMPONENT INDEX
        print("🔎 Verifying component index...")
        if not component_index:
            logging.error("Empty component index - cannot proceed with hyperlinks")
            return False

        if not verify_component_index_integrity(wb):
            logging.error("Component index verification failed")
            return False
        
        # 9. CREATE HYPERLINKS (STEP 3)
        print("🔗 Creating hyperlinks between components...")
        hyperlink_stats = create_hyperlinks(wb)
        
        if not hyperlink_stats['success']:
            logging.error(f"Hyperlink creation completed with {hyperlink_stats['errors']} errors")
            if hyperlink_stats['missing_targets']:
                logging.warning(f"Missing targets: {len(hyperlink_stats['missing_targets'])} unique references")  # Fixed parenthesis
        else:
            print(f"✅ Created {hyperlink_stats['success']} hyperlinks successfully")
        
        # 10. FINAL EXPORT
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = EXCEL_OUTPUT_PATH.parent / f"{EXCEL_OUTPUT_PATH.stem}_{timestamp}{EXCEL_OUTPUT_PATH.suffix}"
        
        print("💾 Saving workbook...")
        wb.save(output_path)
        print(f"🎉 Successfully exported: {output_path}")
        
        return True
        
    except Exception as e:
        logging.critical(f"Tracker failed: {str(e)}", exc_info=True)
        print(f"💥 Critical error: {str(e)}")
        sys.exit(1)
        
if __name__ == "__main__":
    import cProfile
    import pstats

    profile_file = "profile_stats.prof"
    print("🧪 Profiling Laravel Tracker...")
    cProfile.run("main()", profile_file)

    with open("profile_summary.txt", "w") as f:
        stats = pstats.Stats(profile_file, stream=f)
        stats.strip_dirs().sort_stats("tottime").print_stats(40)
    print("📄 Profiling complete → Check 'profile_summary.txt'")
